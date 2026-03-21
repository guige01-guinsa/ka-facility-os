"""Excel complaint import helpers."""

from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domains.complaints import service


_REQUIRED_HEADERS = {"동", "호수", "민원내용"}


@dataclass(frozen=True)
class ImportedComplaintRow:
    site: str
    building: str
    unit_number: str
    description: str
    contact_phone: str | None
    reported_at: datetime | None
    complaint_type: str
    case_key: str
    source_channel: str
    source_workbook: str
    source_sheet: str
    source_row_number: int
    source_row_hash: str


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_text(value))


def _find_header_row(sheet: Any) -> tuple[int | None, dict[str, int]]:
    for row_index in range(1, min(sheet.max_row, 8) + 1):
        headers: dict[str, int] = {}
        for cell in sheet[row_index]:
            normalized = _normalize_header(cell.value)
            if normalized:
                headers[normalized] = cell.column - 1
        if _REQUIRED_HEADERS.issubset(headers):
            return row_index, headers
    return None, {}


def _sheet_default_mapping(sheet_title: str) -> dict[str, int]:
    if sheet_title in {"루버창", "유리"}:
        return {"순번": 0, "날짜": 1, "동": 2, "호수": 3, "민원내용": 4, "전화번호": 5}
    return {"순번": 0, "날짜": 1, "동": 2, "호수": 3, "민원내용": 4, "전화번호": 5}


def _parse_reported_at(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=timezone.utc)
    text = _cell_text(value)
    if not text or text == "추가접수":
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _is_placeholder_row(values: list[Any]) -> bool:
    return any("추가접수" in _cell_text(value) for value in values)


def _sheet_building_hint(sheet_title: str) -> str | None:
    text = _cell_text(sheet_title)
    if text.endswith("동"):
        return text
    return None


def iter_workbook_rows(path: Path, *, site: str, source_channel: str = "legacy_excel") -> list[ImportedComplaintRow]:
    workbook = load_workbook(path, data_only=True)
    imported_rows: list[ImportedComplaintRow] = []
    for sheet in workbook.worksheets:
        header_row, header_map = _find_header_row(sheet)
        row_start = header_row + 1 if header_row is not None else 1
        if not header_map:
            header_map = _sheet_default_mapping(sheet.title)
            while row_start <= sheet.max_row:
                candidate = list(sheet.iter_rows(min_row=row_start, max_row=row_start, values_only=True))[0]
                if any(_cell_text(item) for item in candidate):
                    break
                row_start += 1

        building_hint = _sheet_building_hint(sheet.title)
        for row_index in range(row_start, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))[0]
            if not any(_cell_text(item) for item in row):
                continue
            if _is_placeholder_row(row):
                continue

            building_raw = row[header_map.get("동", 2)] if len(row) > header_map.get("동", 2) else None
            unit_raw = row[header_map.get("호수", 3)] if len(row) > header_map.get("호수", 3) else None
            description_raw = row[header_map.get("민원내용", 4)] if len(row) > header_map.get("민원내용", 4) else None
            phone_raw = row[header_map.get("전화번호", 5)] if len(row) > header_map.get("전화번호", 5) else None
            reported_at_raw = row[header_map.get("날짜", 1)] if len(row) > header_map.get("날짜", 1) else None

            description = service.normalize_description(description_raw)
            if not description:
                continue

            building_source = building_raw if _cell_text(building_raw) else building_hint
            if not building_source or not _cell_text(unit_raw):
                continue

            building = service.normalize_building(building_source)
            unit_number = service.normalize_unit_number(unit_raw)
            contact_phone = service.normalize_phone(phone_raw)
            reported_at = _parse_reported_at(reported_at_raw)
            complaint_type = service.classify_complaint_type(description)
            case_key = service.build_case_key(
                site=site,
                building=building,
                unit_number=unit_number,
                description=description,
                contact_phone=contact_phone,
                reported_at=reported_at,
            )
            source_payload = {
                "workbook": path.name,
                "sheet": sheet.title,
                "row_number": row_index,
                "site": site,
                "building": building,
                "unit_number": unit_number,
                "description": description,
                "contact_phone": contact_phone,
                "reported_at": reported_at.isoformat() if reported_at else None,
            }
            source_row_hash = hashlib.sha256(
                json.dumps(source_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest()
            imported_rows.append(
                ImportedComplaintRow(
                    site=site,
                    building=building,
                    unit_number=unit_number,
                    description=description,
                    contact_phone=contact_phone,
                    reported_at=reported_at,
                    complaint_type=complaint_type,
                    case_key=case_key,
                    source_channel=source_channel,
                    source_workbook=path.name,
                    source_sheet=sheet.title,
                    source_row_number=row_index,
                    source_row_hash=source_row_hash,
                )
            )
    return imported_rows


def dedupe_rows(rows: list[ImportedComplaintRow]) -> tuple[list[ImportedComplaintRow], dict[str, list[ImportedComplaintRow]]]:
    seen: dict[str, ImportedComplaintRow] = {}
    duplicates: dict[str, list[ImportedComplaintRow]] = {}
    for row in rows:
        if row.case_key not in seen:
            seen[row.case_key] = row
            continue
        duplicates.setdefault(row.case_key, []).append(row)
    return list(seen.values()), duplicates


def summarize_workbooks(paths: list[Path], *, site: str) -> dict[str, Any]:
    all_rows: list[ImportedComplaintRow] = []
    for path in paths:
        all_rows.extend(iter_workbook_rows(path, site=site))
    deduped, duplicates = dedupe_rows(all_rows)
    type_counts: dict[str, int] = {}
    workbook_counts: dict[str, int] = {}
    for row in deduped:
        type_counts[row.complaint_type] = type_counts.get(row.complaint_type, 0) + 1
        workbook_counts[row.source_workbook] = workbook_counts.get(row.source_workbook, 0) + 1
    return {
        "site": site,
        "raw_rows": len(all_rows),
        "unique_cases": len(deduped),
        "duplicate_rows": sum(len(items) for items in duplicates.values()),
        "type_counts": dict(sorted(type_counts.items())),
        "workbook_counts": dict(sorted(workbook_counts.items())),
    }
