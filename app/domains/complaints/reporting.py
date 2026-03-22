from __future__ import annotations

import base64
import binascii
import io
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader, simpleSplit
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from sqlalchemy import select

from app.database import (
    complaint_attachments,
    complaint_cases,
    complaint_cost_items,
    complaint_events,
    complaint_messages,
    get_conn,
)
from app.domains.complaints import service
from app.domains.complaints.schemas import ComplaintCaseRead, ComplaintReportCoverOptions


REPORT_TYPE_LABELS: dict[str, str] = {
    "all": "전체",
    "building": "동별",
    "complaint": "민원",
    "category": "분류별",
    "unresolved": "미처리",
    "closed": "종결",
}
ACTIVE_STATUSES = {"received", "assigned", "visit_scheduled", "in_progress", "reopened"}
CLOSED_STATUSES = {"resolved", "resident_confirmed", "closed"}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PDF_FONT_NAME = "HYSMyeongJo-Medium"
_PDF_FONT_READY = False
PDF_MARGIN = 12 * mm
PDF_HEADER_BLUE = colors.HexColor("#1F4E78")
PDF_SOFT_BLUE = colors.HexColor("#EAF2FB")
PDF_LINE = colors.HexColor("#C9D8EA")
PDF_TEXT = colors.HexColor("#18344E")
PDF_MUTED = colors.HexColor("#5C738A")
PDF_OK = colors.HexColor("#1E7E52")
PDF_WARN = colors.HexColor("#D08B1F")
PDF_DANGER = colors.HexColor("#C55A4A")
PDF_CARD_BG = colors.HexColor("#F8FBFF")
PDF_LOGO_GREEN = colors.HexColor("#0C6D58")
PDF_LOGO_ORANGE = colors.HexColor("#CA5F2D")
PDF_TITLE_BG = colors.HexColor("#F3F8FE")


@dataclass(slots=True)
class ComplaintExportReport:
    report_type: str
    report_label: str
    site: str | None
    building: str | None
    generated_at: datetime
    summary_rows: list[tuple[str, str]]
    headers: list[str]
    rows: list[list[str]]
    primary_sheet_name: str
    file_stem: str
    raw_sheets: list[tuple[str, list[str], list[list[str]]]]
    cover_settings: "ComplaintPdfCoverSettings | None" = None


@dataclass(slots=True)
class ComplaintPdfCoverSettings:
    company_name: str | None = None
    contractor_name: str | None = None
    submission_phrase: str | None = None
    logo_bytes: bytes | None = None
    logo_file_name: str | None = None
    logo_content_type: str | None = None


def normalize_report_type(value: str | None) -> str:
    normalized = service.normalize_description(value).lower()
    if normalized not in REPORT_TYPE_LABELS:
        return "all"
    return normalized


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone().strftime("%Y-%m-%d %H:%M")


def _autosize(ws: object) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(text) + 2, 48))
        ws.column_dimensions[letter].width = width


def _style_table_sheet(ws: object) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _normalize_cover_text(value: str | None, *, fallback: str | None = None, max_length: int = 240) -> str | None:
    normalized = service.normalize_description(value)
    if not normalized:
        return fallback
    return normalized[:max_length]


def _decode_logo_data_url(value: str | None) -> tuple[bytes | None, str | None]:
    raw = str(value or "").strip()
    if not raw:
        return None, None
    matched = re.match(r"^data:(image/(png|jpeg|jpg));base64,(.+)$", raw, re.IGNORECASE)
    if matched is None:
        return None, None
    content_type = matched.group(1).lower().replace("jpg", "jpeg")
    try:
        payload = base64.b64decode(matched.group(3), validate=True)
    except (ValueError, binascii.Error):
        return None, None
    if not payload or len(payload) > 1_500_000:
        return None, None
    return payload, content_type


def build_cover_settings(options: ComplaintReportCoverOptions | None) -> ComplaintPdfCoverSettings | None:
    if options is None:
        return None
    logo_bytes, logo_content_type = _decode_logo_data_url(options.logo_data_url)
    settings = ComplaintPdfCoverSettings(
        company_name=_normalize_cover_text(options.company_name, max_length=120),
        contractor_name=_normalize_cover_text(options.contractor_name, max_length=120),
        submission_phrase=_normalize_cover_text(options.submission_phrase, max_length=500),
        logo_bytes=logo_bytes,
        logo_file_name=_normalize_cover_text(options.logo_file_name, max_length=200),
        logo_content_type=logo_content_type,
    )
    if not any(
        (
            settings.company_name,
            settings.contractor_name,
            settings.submission_phrase,
            settings.logo_bytes,
        )
    ):
        return None
    return settings


def _summary_rows(cases: list[ComplaintCaseRead], *, site: str | None, building: str | None, report_label: str) -> list[tuple[str, str]]:
    building_count = len({(row.building, row.unit_number) for row in cases})
    recurrence_count = sum(1 for row in cases if row.recurrence_flag)
    active_count = sum(1 for row in cases if row.status in ACTIVE_STATUSES)
    closed_count = sum(1 for row in cases if row.status in CLOSED_STATUSES)
    latest_report = max((row.reported_at for row in cases), default=None)
    return [
        ("출력구분", report_label),
        ("단지", site or "전체"),
        ("동 필터", building or "전체"),
        ("민원건수", str(len(cases))),
        ("세대수", str(building_count)),
        ("미처리건수", str(active_count)),
        ("종결건수", str(closed_count)),
        ("재민원건수", str(recurrence_count)),
        ("최근접수", _format_datetime(latest_report) or "-"),
    ]


def _detail_rows(cases: Iterable[ComplaintCaseRead]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in cases:
        rows.append(
            [
                str(row.id),
                row.building,
                row.unit_number,
                row.complaint_type_label,
                row.status_label,
                row.assignee or "미배정",
                _format_datetime(row.reported_at),
                row.contact_phone or "",
                row.description,
            ]
        )
    return rows


def _stringify_raw_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.astimezone().isoformat() if value.tzinfo is not None else value.isoformat(sep=" ")
    return str(value)


def _build_raw_sheet(*, name: str, headers: list[str], records: list[dict[str, object]]) -> tuple[str, list[str], list[list[str]]]:
    rows = [[_stringify_raw_value(record.get(header)) for header in headers] for record in records]
    return name, headers, rows


def _load_raw_record_sheets(case_ids: list[int]) -> list[tuple[str, list[str], list[list[str]]]]:
    normalized_case_ids = [int(item) for item in case_ids]
    with get_conn() as conn:
        case_rows = (
            conn.execute(select(complaint_cases).where(complaint_cases.c.id.in_(normalized_case_ids)).order_by(complaint_cases.c.id.asc())).mappings().all()
            if normalized_case_ids
            else []
        )
        event_rows = (
            conn.execute(
                select(complaint_events)
                .where(complaint_events.c.complaint_id.in_(normalized_case_ids))
                .order_by(complaint_events.c.complaint_id.asc(), complaint_events.c.id.asc())
            ).mappings().all()
            if normalized_case_ids
            else []
        )
        attachment_rows = (
            conn.execute(
                select(complaint_attachments)
                .where(complaint_attachments.c.complaint_id.in_(normalized_case_ids))
                .order_by(complaint_attachments.c.complaint_id.asc(), complaint_attachments.c.id.asc())
            ).mappings().all()
            if normalized_case_ids
            else []
        )
        message_rows = (
            conn.execute(
                select(complaint_messages)
                .where(complaint_messages.c.complaint_id.in_(normalized_case_ids))
                .order_by(complaint_messages.c.complaint_id.asc(), complaint_messages.c.id.asc())
            ).mappings().all()
            if normalized_case_ids
            else []
        )
        cost_rows = (
            conn.execute(
                select(complaint_cost_items)
                .where(complaint_cost_items.c.complaint_id.in_(normalized_case_ids))
                .order_by(complaint_cost_items.c.complaint_id.asc(), complaint_cost_items.c.id.asc())
            ).mappings().all()
            if normalized_case_ids
            else []
        )

    attachment_headers = [column.name for column in complaint_attachments.columns if column.name != "file_bytes"] + ["file_bytes_size"]
    attachment_records = []
    for row in attachment_rows:
        item = dict(row)
        file_bytes = item.pop("file_bytes", None)
        item["file_bytes_size"] = len(file_bytes or b"")
        attachment_records.append(item)

    return [
        _build_raw_sheet(
            name="db_complaint_cases",
            headers=[column.name for column in complaint_cases.columns],
            records=[dict(row) for row in case_rows],
        ),
        _build_raw_sheet(
            name="db_complaint_events",
            headers=[column.name for column in complaint_events.columns],
            records=[dict(row) for row in event_rows],
        ),
        _build_raw_sheet(
            name="db_complaint_attachments",
            headers=attachment_headers,
            records=attachment_records,
        ),
        _build_raw_sheet(
            name="db_complaint_messages",
            headers=[column.name for column in complaint_messages.columns],
            records=[dict(row) for row in message_rows],
        ),
        _build_raw_sheet(
            name="db_complaint_cost_items",
            headers=[column.name for column in complaint_cost_items.columns],
            records=[dict(row) for row in cost_rows],
        ),
    ]


def _building_rows(cases: list[ComplaintCaseRead]) -> list[list[str]]:
    grouped: dict[str, dict[str, object]] = {}
    for row in cases:
        bucket = grouped.setdefault(
            row.building,
            {
                "total": 0,
                "active": 0,
                "closed": 0,
                "recurrence": 0,
                "households": set(),
                "latest_reported_at": None,
            },
        )
        bucket["total"] = int(bucket["total"]) + 1
        if row.status in ACTIVE_STATUSES:
            bucket["active"] = int(bucket["active"]) + 1
        if row.status in CLOSED_STATUSES:
            bucket["closed"] = int(bucket["closed"]) + 1
        if row.recurrence_flag:
            bucket["recurrence"] = int(bucket["recurrence"]) + 1
        households = bucket["households"]
        assert isinstance(households, set)
        households.add((row.building, row.unit_number))
        latest_reported_at = bucket["latest_reported_at"]
        if latest_reported_at is None or row.reported_at > latest_reported_at:
            bucket["latest_reported_at"] = row.reported_at

    ordered_keys = sorted(grouped.keys(), key=lambda item: (int("".join(ch for ch in item if ch.isdigit()) or "999999"), item))
    rows: list[list[str]] = []
    for building in ordered_keys:
        bucket = grouped[building]
        households = bucket["households"]
        assert isinstance(households, set)
        rows.append(
            [
                building,
                str(bucket["total"]),
                str(len(households)),
                str(bucket["active"]),
                str(bucket["closed"]),
                str(bucket["recurrence"]),
                _format_datetime(bucket["latest_reported_at"]),
            ]
        )
    return rows


def _category_rows(cases: list[ComplaintCaseRead]) -> list[list[str]]:
    grouped: dict[str, dict[str, object]] = {}
    for row in cases:
        bucket = grouped.setdefault(
            row.complaint_type,
            {
                "label": row.complaint_type_label,
                "total": 0,
                "active": 0,
                "closed": 0,
                "recurrence": 0,
                "households": set(),
            },
        )
        bucket["total"] = int(bucket["total"]) + 1
        if row.status in ACTIVE_STATUSES:
            bucket["active"] = int(bucket["active"]) + 1
        if row.status in CLOSED_STATUSES:
            bucket["closed"] = int(bucket["closed"]) + 1
        if row.recurrence_flag:
            bucket["recurrence"] = int(bucket["recurrence"]) + 1
        households = bucket["households"]
        assert isinstance(households, set)
        households.add((row.building, row.unit_number))

    ordered_keys = sorted(grouped.keys(), key=lambda item: (-int(grouped[item]["total"]), grouped[item]["label"]))  # type: ignore[arg-type]
    rows: list[list[str]] = []
    for complaint_type in ordered_keys:
        bucket = grouped[complaint_type]
        households = bucket["households"]
        assert isinstance(households, set)
        rows.append(
            [
                str(bucket["label"]),
                str(bucket["total"]),
                str(len(households)),
                str(bucket["active"]),
                str(bucket["closed"]),
                str(bucket["recurrence"]),
            ]
        )
    return rows


def build_complaint_export_report(
    *,
    site: str | None,
    report_type: str | None,
    building: str | None = None,
    allowed_sites: list[str] | None = None,
    cover_options: ComplaintReportCoverOptions | None = None,
) -> ComplaintExportReport:
    normalized_type = normalize_report_type(report_type)
    normalized_building = service.normalize_building(building) if service.normalize_description(building) else None
    cases = service.list_cases(site=site, building=normalized_building, allowed_sites=allowed_sites)
    if normalized_type == "unresolved":
        cases = [row for row in cases if row.status in ACTIVE_STATUSES]
    elif normalized_type == "closed":
        cases = [row for row in cases if row.status in CLOSED_STATUSES]
    report_label = REPORT_TYPE_LABELS[normalized_type]
    summary_rows = _summary_rows(cases, site=site, building=normalized_building, report_label=report_label)

    if normalized_type == "building":
        headers = ["동", "총건수", "세대수", "미처리", "종결", "재민원", "최근접수"]
        rows = _building_rows(cases)
        sheet_name = "동별현황"
    elif normalized_type == "category":
        headers = ["분류", "총건수", "세대수", "미처리", "종결", "재민원"]
        rows = _category_rows(cases)
        sheet_name = "분류별현황"
    else:
        headers = ["민원ID", "동", "호수", "민원유형", "상태", "담당자", "접수일시", "연락처", "민원내용"]
        rows = _detail_rows(cases)
        sheet_name = "민원목록"

    file_stem = f"complaints-{normalized_type}-{service.normalize_description(site or 'all').replace(' ', '_') or 'all'}"
    raw_sheets = _load_raw_record_sheets([row.id for row in cases])
    return ComplaintExportReport(
        report_type=normalized_type,
        report_label=report_label,
        site=site,
        building=normalized_building,
        generated_at=datetime.now().astimezone(),
        summary_rows=summary_rows,
        headers=headers,
        rows=rows,
        primary_sheet_name=sheet_name,
        file_stem=file_stem,
        raw_sheets=raw_sheets,
        cover_settings=build_cover_settings(cover_options),
    )


def build_complaint_export_xlsx(report: ComplaintExportReport) -> bytes:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "요약"
    summary_ws.append(["항목", "값"])
    for label, value in report.summary_rows:
        summary_ws.append([label, value])
    _style_table_sheet(summary_ws)

    detail_ws = workbook.create_sheet(report.primary_sheet_name)
    detail_ws.append(report.headers)
    for row in report.rows:
        detail_ws.append(row)
    _style_table_sheet(detail_ws)

    for sheet_name, headers, rows in report.raw_sheets:
        raw_ws = workbook.create_sheet(sheet_name)
        raw_ws.append(headers)
        for row in rows:
            raw_ws.append(row)
        _style_table_sheet(raw_ws)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _ensure_pdf_font() -> str:
    global _PDF_FONT_READY
    if not _PDF_FONT_READY:
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(PDF_FONT_NAME))
        except Exception:
            return "Helvetica"
        _PDF_FONT_READY = True
    return PDF_FONT_NAME


def _summary_value_map(report: ComplaintExportReport) -> dict[str, str]:
    return {label: value for label, value in report.summary_rows}


def _int_from_summary(value: str | None) -> int:
    text = str(value or "").strip()
    if not text:
        return 0
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return 0
    return int(digits)


def _raw_case_records(report: ComplaintExportReport) -> list[dict[str, str]]:
    for sheet_name, headers, rows in report.raw_sheets:
        if sheet_name != "db_complaint_cases":
            continue
        return [dict(zip(headers, row)) for row in rows]
    return []


def _progress_metrics(report: ComplaintExportReport) -> dict[str, float]:
    raw_records = _raw_case_records(report)
    summary = _summary_value_map(report)
    total_count = _int_from_summary(summary.get("민원건수"))
    if raw_records:
        status_counts = Counter(str(row.get("status") or "") for row in raw_records)
        received_count = int(status_counts.get("received", 0))
        started_count = max(total_count - received_count, 0)
        closed_count = sum(int(status_counts.get(status, 0)) for status in CLOSED_STATUSES)
    else:
        closed_count = _int_from_summary(summary.get("종결건수"))
        started_count = max(total_count - _int_from_summary(summary.get("미처리건수")), closed_count)
    recurrence_count = _int_from_summary(summary.get("재민원건수"))
    progress_rate = (started_count / total_count) if total_count else 0.0
    completion_rate = (closed_count / total_count) if total_count else 0.0
    return {
        "total": float(total_count),
        "started": float(started_count),
        "closed": float(closed_count),
        "recurrence": float(recurrence_count),
        "progress_rate": progress_rate,
        "completion_rate": completion_rate,
    }


def _cover_metadata(report: ComplaintExportReport) -> dict[str, str]:
    cover = report.cover_settings
    site_label = service.normalize_description(report.site or "전체 단지")
    building_scope = service.normalize_description(report.building or "전체 동")
    report_date = report.generated_at.strftime("%Y년 %m월 %d일")
    return {
        "submission_label": "관리사무소 제출용",
        "document_title": "세대 민원 처리 현황 보고서",
        "document_subtitle": "외벽 도색 작업 중 접수된 세대 민원과 처리 진행 현황을 정리한 제출 문서",
        "recipient": f"{site_label} 관리사무소",
        "site_name": site_label,
        "construction_name": f"{site_label} 외벽 재도장 공사 세대 민원 처리",
        "report_date": report_date,
        "scope": building_scope,
        "report_kind": f"{report.report_label} 보고",
        "company_name": _normalize_cover_text(getattr(cover, "company_name", None), fallback="KA Facility OS", max_length=120) or "KA Facility OS",
        "contractor_name": _normalize_cover_text(getattr(cover, "contractor_name", None), fallback="외벽 재도장 협력업체", max_length=120) or "외벽 재도장 협력업체",
        "company_subtitle": "Field Operations Reporting",
        "submission_copy": _normalize_cover_text(getattr(cover, "submission_phrase", None), fallback="상기 현황을 아래와 같이 보고드립니다.", max_length=500) or "상기 현황을 아래와 같이 보고드립니다.",
    }


def _draw_pdf_frame(pdf: canvas.Canvas, *, width: float, height: float) -> None:
    pdf.setStrokeColor(PDF_LINE)
    pdf.setLineWidth(1)
    pdf.roundRect(PDF_MARGIN, PDF_MARGIN, width - 2 * PDF_MARGIN, height - 2 * PDF_MARGIN, 8, stroke=1, fill=0)


def _draw_approval_box(pdf: canvas.Canvas, *, x: float, y_top: float, width: float, height: float, font_name: str) -> None:
    roles = ("주임", "계장", "과장", "소장")
    label_height = 8 * mm
    title_width = 14 * mm
    cell_width = (width - title_width) / len(roles)
    y_bottom = y_top - height
    pdf.setStrokeColor(PDF_LINE)
    pdf.setFillColor(colors.white)
    pdf.roundRect(x, y_bottom, width, height, 6, stroke=1, fill=1)
    pdf.setFillColor(PDF_SOFT_BLUE)
    pdf.rect(x, y_top - label_height, width, label_height, stroke=0, fill=1)
    pdf.setFillColor(colors.white)
    pdf.rect(x, y_bottom, title_width, height, stroke=0, fill=1)
    pdf.line(x, y_top - label_height, x + width, y_top - label_height)
    pdf.line(x + title_width, y_bottom, x + title_width, y_top)
    for index in range(1, len(roles)):
        current_x = x + title_width + cell_width * index
        pdf.line(current_x, y_bottom, current_x, y_top)
    pdf.setFont(font_name, 9)
    pdf.setFillColor(PDF_TEXT)
    pdf.drawCentredString(x + title_width / 2, y_bottom + height / 2 - 3, "결재")
    pdf.setFont(font_name, 8)
    pdf.setFillColor(PDF_MUTED)
    for index, role in enumerate(roles):
        center_x = x + title_width + cell_width * index + cell_width / 2
        pdf.drawCentredString(center_x, y_top - 5.5 * mm, role)
        pdf.setStrokeColor(PDF_LINE)
        pdf.line(
            x + title_width + cell_width * index + 4,
            y_bottom + 9,
            x + title_width + cell_width * (index + 1) - 4,
            y_bottom + 9,
        )


def _draw_company_logo(
    pdf: canvas.Canvas,
    *,
    x: float,
    y_top: float,
    width: float,
    height: float,
    font_name: str,
    metadata: dict[str, str],
    report: ComplaintExportReport,
) -> None:
    y_bottom = y_top - height
    pdf.setStrokeColor(PDF_LINE)
    pdf.setFillColor(colors.white)
    pdf.roundRect(x, y_bottom, width, height, 8, stroke=1, fill=1)
    icon_size = min(height - 10, 18 * mm)
    icon_x = x + 7
    icon_y = y_bottom + (height - icon_size) / 2
    cover = report.cover_settings
    image_drawn = False
    if cover and cover.logo_bytes:
        try:
            reader = ImageReader(io.BytesIO(cover.logo_bytes))
            pdf.drawImage(
                reader,
                icon_x,
                icon_y,
                width=icon_size,
                height=icon_size,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )
            image_drawn = True
        except Exception:
            image_drawn = False
    if not image_drawn:
        pdf.setFillColor(PDF_LOGO_GREEN)
        pdf.roundRect(icon_x, icon_y, icon_size, icon_size, 5, stroke=0, fill=1)
        pdf.setFillColor(PDF_LOGO_ORANGE)
        pdf.circle(icon_x + icon_size - 4, icon_y + icon_size - 4, 4, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont(font_name, 14)
        pdf.drawCentredString(icon_x + icon_size / 2, icon_y + icon_size / 2 - 5, "KA")
    text_x = icon_x + icon_size + 8
    pdf.setFillColor(PDF_TEXT)
    pdf.setFont(font_name, 11)
    pdf.drawString(text_x, y_bottom + height - 14, metadata["company_name"])
    pdf.setFillColor(PDF_MUTED)
    pdf.setFont(font_name, 8)
    pdf.drawString(text_x, y_bottom + height - 25, metadata["company_subtitle"])
    contractor_line = simpleSplit("시공사 " + metadata["contractor_name"], font_name, 8, max(width - icon_size - 20, 40))
    pdf.drawString(text_x, y_bottom + 9, (contractor_line or ["관리사무소 제출 문서"])[0][:48])


def _draw_cover_information_panel(
    pdf: canvas.Canvas,
    report: ComplaintExportReport,
    *,
    width: float,
    start_y: float,
    font_name: str,
    metadata: dict[str, str],
) -> float:
    panel_height = 30 * mm
    y_bottom = start_y - panel_height
    pdf.setStrokeColor(PDF_LINE)
    pdf.setFillColor(colors.white)
    pdf.roundRect(PDF_MARGIN, y_bottom, width - 2 * PDF_MARGIN, panel_height, 8, stroke=1, fill=1)
    pdf.setFont(font_name, 8)
    pdf.setFillColor(PDF_MUTED)
    row_one = [
        ("제출처", metadata["recipient"]),
        ("보고일", metadata["report_date"]),
        ("문서구분", metadata["report_kind"]),
    ]
    row_two = [
        ("단지명", metadata["site_name"]),
        ("제출사", metadata["company_name"]),
        ("시공사", metadata["contractor_name"]),
    ]
    row_three = [("공사명", metadata["construction_name"])]
    cell_width = (width - 2 * PDF_MARGIN - 16) / 3
    row_positions = [y_bottom + panel_height - 12, y_bottom + panel_height - 25]
    for row_index, row_items in enumerate((row_one, row_two)):
        current_x = PDF_MARGIN + 8
        current_y = row_positions[row_index]
        for label, value in row_items:
            pdf.setFillColor(PDF_MUTED)
            pdf.drawString(current_x, current_y, f"{label}")
            pdf.setFillColor(PDF_TEXT)
            pdf.drawString(current_x + 17 * mm, current_y, str(value)[:36])
            current_x += cell_width
    pdf.setFillColor(PDF_MUTED)
    pdf.drawString(PDF_MARGIN + 8, y_bottom + 14, row_three[0][0])
    pdf.setFillColor(PDF_TEXT)
    wrapped = simpleSplit(row_three[0][1], font_name, 8, width - 2 * PDF_MARGIN - 30 * mm)
    for index, line in enumerate((wrapped or [row_three[0][1]])[:2]):
        pdf.drawString(PDF_MARGIN + 25 * mm, y_bottom + 14 - index * 9, line)
    return y_bottom - 8


def _draw_summary_cards(pdf: canvas.Canvas, report: ComplaintExportReport, *, width: float, start_y: float, font_name: str) -> float:
    summary = _summary_value_map(report)
    cards = [
        ("민원건수", summary.get("민원건수", "0"), PDF_HEADER_BLUE),
        ("세대수", summary.get("세대수", "0"), colors.HexColor("#2D6F92")),
        ("미처리", summary.get("미처리건수", "0"), PDF_WARN),
        ("종결", summary.get("종결건수", "0"), PDF_OK),
        ("재민원", summary.get("재민원건수", "0"), PDF_DANGER),
    ]
    gap = 5 * mm
    card_width = (width - 2 * PDF_MARGIN - gap * 4) / 5
    card_height = 23 * mm
    y_bottom = start_y - card_height
    for index, (label, value, accent) in enumerate(cards):
        x = PDF_MARGIN + index * (card_width + gap)
        pdf.setStrokeColor(PDF_LINE)
        pdf.setFillColor(PDF_CARD_BG)
        pdf.roundRect(x, y_bottom, card_width, card_height, 8, stroke=1, fill=1)
        pdf.setFillColor(accent)
        pdf.roundRect(x + 4, y_bottom + card_height - 7, card_width - 8, 3, 2, stroke=0, fill=1)
        pdf.setFillColor(PDF_MUTED)
        pdf.setFont(font_name, 8)
        pdf.drawString(x + 7, y_bottom + card_height - 16, label)
        pdf.setFillColor(PDF_TEXT)
        pdf.setFont(font_name, 15)
        pdf.drawString(x + 7, y_bottom + 9, str(value))
    pdf.setFillColor(PDF_MUTED)
    pdf.setFont(font_name, 8)
    pdf.drawString(
        PDF_MARGIN,
        y_bottom - 11,
        f"보고기준 {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')} · 단지 {report.site or '전체'} · 범위 {report.building or '전체'}",
    )
    return y_bottom - 18


def _draw_progress_bars(pdf: canvas.Canvas, report: ComplaintExportReport, *, width: float, start_y: float, font_name: str) -> float:
    metrics = _progress_metrics(report)
    total = max(metrics["total"], 1.0)
    bars = [
        ("진행률", metrics["progress_rate"], f"{int(metrics['started'])}/{int(total)}", PDF_WARN),
        ("완성률", metrics["completion_rate"], f"{int(metrics['closed'])}/{int(total)}", PDF_OK),
    ]
    box_gap = 8 * mm
    box_width = (width - 2 * PDF_MARGIN - box_gap) / 2
    box_height = 24 * mm
    y_bottom = start_y - box_height
    for index, (label, rate, subtitle, accent) in enumerate(bars):
        x = PDF_MARGIN + index * (box_width + box_gap)
        pdf.setStrokeColor(PDF_LINE)
        pdf.setFillColor(colors.white)
        pdf.roundRect(x, y_bottom, box_width, box_height, 8, stroke=1, fill=1)
        pdf.setFillColor(PDF_MUTED)
        pdf.setFont(font_name, 9)
        pdf.drawString(x + 8, y_bottom + box_height - 12, label)
        pdf.setFillColor(PDF_TEXT)
        pdf.setFont(font_name, 13)
        pdf.drawString(x + 8, y_bottom + box_height - 25, f"{int(rate * 100)}%")
        pdf.setFont(font_name, 8)
        pdf.setFillColor(PDF_MUTED)
        pdf.drawString(x + 42, y_bottom + box_height - 24, subtitle)
        bar_x = x + 8
        bar_y = y_bottom + 8
        bar_width = box_width - 16
        pdf.setFillColor(PDF_SOFT_BLUE)
        pdf.roundRect(bar_x, bar_y, bar_width, 8, 4, stroke=0, fill=1)
        pdf.setFillColor(accent)
        pdf.roundRect(bar_x, bar_y, bar_width * max(0.0, min(rate, 1.0)), 8, 4, stroke=0, fill=1)
    return y_bottom - 10


def _draw_status_chart(pdf: canvas.Canvas, report: ComplaintExportReport, *, width: float, start_y: float, font_name: str) -> float:
    raw_records = _raw_case_records(report)
    status_counts = Counter(str(row.get("status") or "") for row in raw_records)
    ordered_statuses = [status for status in ("received", "assigned", "visit_scheduled", "in_progress", "resolved", "resident_confirmed", "reopened", "closed") if status_counts.get(status)]
    if not ordered_statuses:
        ordered_statuses = ["received"]
        status_counts["received"] = 0
    max_count = max(int(status_counts.get(status, 0)) for status in ordered_statuses) or 1
    chart_height = 10 * mm + len(ordered_statuses) * 8 * mm
    y_bottom = start_y - chart_height
    pdf.setStrokeColor(PDF_LINE)
    pdf.setFillColor(colors.white)
    pdf.roundRect(PDF_MARGIN, y_bottom, width - 2 * PDF_MARGIN, chart_height, 8, stroke=1, fill=1)
    pdf.setFillColor(PDF_TEXT)
    pdf.setFont(font_name, 10)
    pdf.drawString(PDF_MARGIN + 8, start_y - 12, "상태 분포")
    bar_origin_x = PDF_MARGIN + 45 * mm
    bar_max_width = width - PDF_MARGIN - bar_origin_x - 14
    current_y = start_y - 24
    for status in ordered_statuses:
        label = service.STATUS_LABELS.get(status, status)
        count = int(status_counts.get(status, 0))
        if status in CLOSED_STATUSES:
            accent = PDF_OK
        elif status == "received":
            accent = PDF_DANGER
        else:
            accent = PDF_WARN
        pdf.setFont(font_name, 8)
        pdf.setFillColor(PDF_MUTED)
        pdf.drawString(PDF_MARGIN + 8, current_y + 2, label)
        pdf.setFillColor(PDF_SOFT_BLUE)
        pdf.roundRect(bar_origin_x, current_y, bar_max_width, 5, 2, stroke=0, fill=1)
        pdf.setFillColor(accent)
        pdf.roundRect(bar_origin_x, current_y, bar_max_width * (count / max_count), 5, 2, stroke=0, fill=1)
        pdf.setFillColor(PDF_TEXT)
        pdf.drawRightString(width - PDF_MARGIN - 8, current_y + 1.5, str(count))
        current_y -= 8 * mm
    return y_bottom - 10


def _pdf_table_layout(report: ComplaintExportReport) -> tuple[list[str], list[float], float]:
    if report.report_type == "building":
        return report.headers, [17 * mm, 17 * mm, 17 * mm, 17 * mm, 17 * mm, 17 * mm, 44 * mm], 7.6
    if report.report_type == "category":
        return report.headers, [32 * mm, 18 * mm, 18 * mm, 18 * mm, 18 * mm, 42 * mm], 7.8
    return report.headers, [10 * mm, 12 * mm, 12 * mm, 18 * mm, 14 * mm, 16 * mm, 20 * mm, 20 * mm, 53 * mm], 7.0


def _draw_table_header(pdf: canvas.Canvas, *, headers: list[str], widths: list[float], x: float, y_top: float, font_name: str, font_size: float) -> float:
    header_height = 14
    pdf.setFillColor(PDF_HEADER_BLUE)
    pdf.setStrokeColor(PDF_HEADER_BLUE)
    pdf.rect(x, y_top - header_height, sum(widths), header_height, stroke=1, fill=1)
    pdf.setFillColor(colors.white)
    pdf.setFont(font_name, font_size)
    current_x = x
    for header, width in zip(headers, widths):
        pdf.drawString(current_x + 4, y_top - 10, header[:20])
        current_x += width
    return y_top - header_height


def _draw_table_rows(
    pdf: canvas.Canvas,
    *,
    headers: list[str],
    widths: list[float],
    rows: list[list[str]],
    start_index: int,
    x: float,
    y_top: float,
    bottom_y: float,
    font_name: str,
    font_size: float,
) -> int:
    y = _draw_table_header(pdf, headers=headers, widths=widths, x=x, y_top=y_top, font_name=font_name, font_size=font_size)
    line_height = font_size + 2
    index = start_index
    while index < len(rows):
        wrapped_cells: list[list[str]] = []
        max_lines = 1
        for cell_value, width in zip(rows[index], widths):
            wrapped = simpleSplit(str(cell_value or ""), font_name, font_size, max(width - 6, 20))
            wrapped_cells.append(wrapped or [""])
            max_lines = max(max_lines, len(wrapped or [""]))
        row_height = max(12, max_lines * line_height + 4)
        if y - row_height < bottom_y:
            return index
        pdf.setStrokeColor(PDF_LINE)
        pdf.setFillColor(colors.white)
        pdf.rect(x, y - row_height, sum(widths), row_height, stroke=1, fill=1)
        current_x = x
        pdf.setFont(font_name, font_size)
        pdf.setFillColor(PDF_TEXT)
        for column_index, width in enumerate(widths):
            if column_index > 0:
                pdf.line(current_x, y, current_x, y - row_height)
            current_y = y - 9
            for line in wrapped_cells[column_index][: max_lines]:
                pdf.drawString(current_x + 3, current_y, line[:120])
                current_y -= line_height
            current_x += width
        y -= row_height
        index += 1
    return index


def _draw_page_footer(pdf: canvas.Canvas, *, width: float, page_number: int, font_name: str) -> None:
    pdf.setFillColor(PDF_MUTED)
    pdf.setFont(font_name, 8)
    pdf.drawString(PDF_MARGIN, PDF_MARGIN - 2, "KA Facility OS · 세대 민원관리 보고서")
    pdf.drawRightString(width - PDF_MARGIN, PDF_MARGIN - 2, f"{page_number} page")


def _draw_cover_page(pdf: canvas.Canvas, report: ComplaintExportReport, *, width: float, height: float, font_name: str) -> float:
    metadata = _cover_metadata(report)
    _draw_pdf_frame(pdf, width=width, height=height)
    approval_width = 96 * mm
    approval_height = 32 * mm
    logo_width = 72 * mm
    logo_height = 28 * mm
    top_y = height - PDF_MARGIN + 2
    _draw_company_logo(
        pdf,
        x=PDF_MARGIN,
        y_top=top_y,
        width=logo_width,
        height=logo_height,
        font_name=font_name,
        metadata=metadata,
        report=report,
    )
    approval_x = width - PDF_MARGIN - approval_width
    _draw_approval_box(pdf, x=approval_x, y_top=top_y, width=approval_width, height=approval_height, font_name=font_name)
    badge_y = height - PDF_MARGIN - 38
    pdf.setFillColor(PDF_TITLE_BG)
    pdf.setStrokeColor(PDF_LINE)
    badge_width = 34 * mm
    badge_height = 8 * mm
    pdf.roundRect(PDF_MARGIN, badge_y - badge_height + 2, badge_width, badge_height, 4, stroke=1, fill=1)
    pdf.setFillColor(PDF_HEADER_BLUE)
    pdf.setFont(font_name, 8)
    pdf.drawCentredString(PDF_MARGIN + badge_width / 2, badge_y - 3.5, metadata["submission_label"])
    pdf.setFillColor(PDF_HEADER_BLUE)
    pdf.setFont(font_name, 18)
    pdf.drawString(PDF_MARGIN, height - PDF_MARGIN - 52, metadata["document_title"])
    pdf.setFont(font_name, 9)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawString(PDF_MARGIN, height - PDF_MARGIN - 64, metadata["document_subtitle"])
    pdf.setFont(font_name, 9)
    pdf.drawString(PDF_MARGIN, height - PDF_MARGIN - 78, metadata["submission_copy"])
    latest_value = _summary_value_map(report).get("최근접수", "-")
    pdf.drawRightString(width - PDF_MARGIN, height - PDF_MARGIN - 78, f"최근 접수 {latest_value}")
    current_y = _draw_cover_information_panel(
        pdf,
        report,
        width=width,
        start_y=height - PDF_MARGIN - 84,
        font_name=font_name,
        metadata=metadata,
    )
    current_y = _draw_summary_cards(pdf, report, width=width, start_y=current_y, font_name=font_name)
    current_y = _draw_progress_bars(pdf, report, width=width, start_y=current_y, font_name=font_name)
    current_y = _draw_status_chart(pdf, report, width=width, start_y=current_y, font_name=font_name)
    pdf.setFillColor(PDF_MUTED)
    pdf.setFont(font_name, 8)
    pdf.drawString(PDF_MARGIN, current_y, "상세 목록")
    return current_y - 6


def build_complaint_export_pdf(report: ComplaintExportReport) -> bytes:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    font_name = _ensure_pdf_font()
    headers, widths, font_size = _pdf_table_layout(report)
    page_number = 1
    next_y = _draw_cover_page(pdf, report, width=width, height=height, font_name=font_name)
    next_index = _draw_table_rows(
        pdf,
        headers=headers,
        widths=widths,
        rows=report.rows,
        start_index=0,
        x=PDF_MARGIN,
        y_top=next_y,
        bottom_y=PDF_MARGIN + 12,
        font_name=font_name,
        font_size=font_size,
    )
    _draw_page_footer(pdf, width=width, page_number=page_number, font_name=font_name)
    while next_index < len(report.rows):
        pdf.showPage()
        page_number += 1
        _draw_pdf_frame(pdf, width=width, height=height)
        pdf.setFillColor(PDF_HEADER_BLUE)
        pdf.setFont(font_name, 14)
        pdf.drawString(PDF_MARGIN, height - PDF_MARGIN - 4, f"세대 민원관리 {report.report_label} 상세목록")
        pdf.setFillColor(PDF_MUTED)
        pdf.setFont(font_name, 8)
        pdf.drawRightString(width - PDF_MARGIN, height - PDF_MARGIN - 4, f"{report.site or '전체'} · {report.building or '전체'}")
        next_index = _draw_table_rows(
            pdf,
            headers=headers,
            widths=widths,
            rows=report.rows,
            start_index=next_index,
            x=PDF_MARGIN,
            y_top=height - PDF_MARGIN - 16,
            bottom_y=PDF_MARGIN + 12,
            font_name=font_name,
            font_size=font_size,
        )
        _draw_page_footer(pdf, width=width, page_number=page_number, font_name=font_name)
    pdf.save()
    return buffer.getvalue()
