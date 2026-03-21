from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.domains.complaints.schemas import ComplaintCaseUpdate, ComplaintEventCreate
from app.domains.complaints.service import STATUS_LABELS, add_event, get_case_detail, update_case


STATUS_ALIASES = {code.lower(): code for code in STATUS_LABELS}
STATUS_ALIASES.update({label.lower(): code for code, label in STATUS_LABELS.items()})
TRUE_VALUES = {"y", "yes", "true", "1", "예"}
FALSE_VALUES = {"n", "no", "false", "0", "아니오"}
CLEAR_VALUES = {"clear", "삭제", "none", "null", "-"}
KST = timezone(timedelta(hours=9))


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply complaint bulk update workbook")
    parser.add_argument("--workbook", required=True, help="xlsx workbook path exported from the bulk update template script")
    parser.add_argument("--apply", action="store_true", help="write updates to the current database")
    parser.add_argument("--actor", default="bulk-template", help="actor username to record in audit/event logs")
    return parser.parse_args()


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value).strip()


def _parse_bool_flag(value: Any) -> bool:
    return _normalize_text(value).lower() in TRUE_VALUES


def _parse_status(value: Any) -> str | None:
    raw = _normalize_text(value).lower()
    if not raw:
        return None
    normalized = STATUS_ALIASES.get(raw)
    if normalized is None:
        raise ValueError(f"unsupported status: {value}")
    return normalized


def _parse_recurrence(value: Any) -> bool | None:
    raw = _normalize_text(value).lower()
    if not raw:
        return None
    if raw in TRUE_VALUES:
        return True
    if raw in FALSE_VALUES:
        return False
    raise ValueError(f"unsupported recurrence flag: {value}")


def _parse_visit_at(value: Any) -> tuple[bool, Any]:
    if isinstance(value, datetime):
        normalized = value.replace(tzinfo=KST) if value.tzinfo is None else value
        return True, normalized.astimezone(timezone.utc)
    raw = _normalize_text(value)
    if not raw:
        return False, None
    if raw.lower() in CLEAR_VALUES:
        return True, None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
        try:
            return True, datetime.strptime(raw, fmt).replace(tzinfo=KST).astimezone(timezone.utc)
        except ValueError:
            continue
    raise ValueError(f"unsupported scheduled visit format: {value}")


def _parse_assignee(value: Any) -> tuple[bool, str | None]:
    raw = _normalize_text(value)
    if not raw:
        return False, None
    if raw.lower() in CLEAR_VALUES:
        return True, ""
    return True, raw


def _principal(actor: str) -> dict[str, Any]:
    return {"username": actor}


def _load_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "업데이트양식" not in wb.sheetnames:
        raise SystemExit("업데이트양식 시트를 찾을 수 없습니다.")
    ws = wb["업데이트양식"]
    headers = [_normalize_text(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None or _normalize_text(value) == "" for value in row):
            continue
        item = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        rows.append(item)
    return rows


def _build_update_payload(row: dict[str, Any], current: Any) -> tuple[dict[str, Any], str]:
    changes: dict[str, Any] = {}
    note = _normalize_text(row.get("처리메모"))
    next_status = _parse_status(row.get("변경상태코드"))
    if next_status and next_status != str(current.case.status):
        changes["status"] = next_status

    assignee_provided, assignee_value = _parse_assignee(row.get("변경담당자"))
    if assignee_provided:
        current_assignee = _normalize_text(current.case.assignee)
        next_assignee = "" if assignee_value is None else str(assignee_value)
        if next_assignee != current_assignee:
            changes["assignee"] = assignee_value

    visit_provided, visit_value = _parse_visit_at(row.get("변경방문예정"))
    if visit_provided:
        current_visit = current.case.scheduled_visit_at
        current_visit_text = current_visit.strftime("%Y-%m-%d %H:%M") if current_visit else ""
        next_visit_text = visit_value.strftime("%Y-%m-%d %H:%M") if isinstance(visit_value, datetime) else ""
        if next_visit_text != current_visit_text:
            changes["scheduled_visit_at"] = visit_value

    recurrence_value = _parse_recurrence(row.get("재민원표시"))
    if recurrence_value is not None and bool(current.case.recurrence_flag) != recurrence_value:
        changes["recurrence_flag"] = recurrence_value

    return changes, note


def process_workbook(*, workbook_path: Path, actor: str, apply: bool) -> dict[str, Any]:
    rows = _load_rows(workbook_path)
    summary: dict[str, Any] = {
        "workbook": str(workbook_path),
        "rows_in_sheet": len(rows),
        "apply_requested": bool(apply),
        "selected_rows": 0,
        "updated_cases": 0,
        "noted_cases": 0,
        "no_change_rows": 0,
        "error_rows": 0,
        "errors": [],
    }
    for row_number, row in enumerate(rows, start=2):
        if not _parse_bool_flag(row.get("적용여부(Y)")):
            continue
        summary["selected_rows"] += 1
        complaint_id_raw = row.get("민원ID")
        try:
            complaint_id = int(complaint_id_raw)
            current = get_case_detail(complaint_id=complaint_id)
            changes, note = _build_update_payload(row, current)
            if not changes and not note:
                summary["no_change_rows"] += 1
                continue
            if not apply:
                if changes:
                    summary["updated_cases"] += 1
                if note:
                    summary["noted_cases"] += 1
                continue
            principal = _principal(actor)
            if changes:
                update_case(
                    complaint_id=complaint_id,
                    payload=ComplaintCaseUpdate(**changes),
                    principal=principal,
                )
                summary["updated_cases"] += 1
            if note:
                add_event(
                    complaint_id=complaint_id,
                    payload=ComplaintEventCreate(
                        event_type="bulk_note",
                        note=note,
                        detail={
                            "source": "bulk_update_template",
                            "workbook": workbook_path.name,
                            "row_number": row_number,
                        },
                    ),
                    principal=principal,
                )
                summary["noted_cases"] += 1
        except Exception as exc:
            summary["error_rows"] += 1
            if len(summary["errors"]) < 20:
                summary["errors"].append(
                    {
                        "row_number": row_number,
                        "complaint_id": complaint_id_raw,
                        "error": str(exc),
                    }
                )
    return summary


def main() -> None:
    args = _parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"workbook not found: {workbook_path}")
    result = process_workbook(workbook_path=workbook_path, actor=args.actor, apply=bool(args.apply))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
