from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.database import complaint_cases, get_conn
from app.domains.complaints.service import PRIORITY_LABELS, STATUS_LABELS, complaint_type_label


KST = timezone(timedelta(hours=9))
ACTIVE_STATUSES = {"received", "assigned", "visit_scheduled", "in_progress", "reopened"}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export complaint report workbook")
    parser.add_argument("--site", required=True, help="site name, for example 연산더샵")
    parser.add_argument("--output", required=True, help="xlsx output path")
    return parser.parse_args()


def _sort_key(building: Any, unit_number: Any) -> tuple[Any, ...]:
    def _numeric_token(value: Any) -> tuple[int, str]:
        text = str(value or "").strip()
        digits = "".join(ch for ch in text if ch.isdigit())
        return (int(digits), text) if digits else (10**9, text)

    return (*_numeric_token(building), *_numeric_token(unit_number))


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_datetime(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.strftime("%Y-%m-%d %H:%M")
        return value.astimezone(KST).strftime("%Y-%m-%d %H:%M")
    return _to_text(value)


def _bool_label(value: Any) -> str:
    return "예" if bool(value) else "아니오"


def _autosize(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(text) + 2, 48))
        ws.column_dimensions[letter].width = width


def _style_sheet(ws: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _append_sheet(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_sheet(ws)


def _load_cases(site: str) -> list[dict[str, Any]]:
    with get_conn() as conn:
        rows = conn.execute(
            select(complaint_cases)
            .where(complaint_cases.c.site == site)
            .order_by(complaint_cases.c.building.asc(), complaint_cases.c.unit_number.asc(), complaint_cases.c.reported_at.asc())
        ).mappings()
        return [dict(row) for row in rows]


def _build_household_rows(cases: list[dict[str, Any]]) -> list[list[Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in cases:
        key = (_to_text(row.get("building")), _to_text(row.get("unit_number")))
        current = grouped.setdefault(
            key,
            {
                "building": key[0],
                "unit_number": key[1],
                "total_count": 0,
                "open_count": 0,
                "recurrence_marked_count": 0,
                "first_reported_at": None,
                "latest_reported_at": None,
                "latest_row": None,
                "types": set(),
            },
        )
        current["total_count"] += 1
        if _to_text(row.get("status")) in ACTIVE_STATUSES:
            current["open_count"] += 1
        if bool(row.get("recurrence_flag")) or int(row.get("recurrence_count") or 0) > 0:
            current["recurrence_marked_count"] += 1
        current["types"].add(complaint_type_label(_to_text(row.get("complaint_type"))))

        reported_at = row.get("reported_at") or row.get("created_at")
        first_reported_at = current["first_reported_at"]
        latest_reported_at = current["latest_reported_at"]
        if first_reported_at is None or (reported_at is not None and reported_at < first_reported_at):
            current["first_reported_at"] = reported_at
        if latest_reported_at is None or (reported_at is not None and reported_at >= latest_reported_at):
            current["latest_reported_at"] = reported_at
            current["latest_row"] = row

    rows: list[list[Any]] = []
    for item in sorted(grouped.values(), key=lambda row: _sort_key(row["building"], row["unit_number"])):
        latest_row = item["latest_row"] or {}
        rows.append(
            [
                item["building"],
                item["unit_number"],
                item["total_count"],
                item["open_count"],
                item["recurrence_marked_count"],
                STATUS_LABELS.get(_to_text(latest_row.get("status")), _to_text(latest_row.get("status"))),
                _format_datetime(item["first_reported_at"]),
                _format_datetime(item["latest_reported_at"]),
                complaint_type_label(_to_text(latest_row.get("complaint_type"))),
                _to_text(latest_row.get("title")),
                _to_text(latest_row.get("assignee")) or "미배정",
                _to_text(latest_row.get("contact_phone")),
                ", ".join(sorted(item["types"])),
                _to_text(latest_row.get("source_sheet")),
                _to_text(latest_row.get("source_row_number")),
            ]
        )
    return rows


def _build_case_rows(cases: list[dict[str, Any]]) -> list[list[Any]]:
    ordered = sorted(
        cases,
        key=lambda row: (
            _sort_key(row.get("building"), row.get("unit_number")),
            row.get("reported_at") or row.get("created_at") or datetime.min.replace(tzinfo=timezone.utc),
            int(row.get("id") or 0),
        ),
    )
    rows: list[list[Any]] = []
    for row in ordered:
        rows.append(
            [
                row.get("id"),
                STATUS_LABELS.get(_to_text(row.get("status")), _to_text(row.get("status"))),
                PRIORITY_LABELS.get(_to_text(row.get("priority")), _to_text(row.get("priority"))),
                _format_datetime(row.get("reported_at")),
                _to_text(row.get("building")),
                _to_text(row.get("unit_number")),
                complaint_type_label(_to_text(row.get("complaint_type"))),
                _to_text(row.get("title")),
                _to_text(row.get("description")),
                _to_text(row.get("resident_name")),
                _to_text(row.get("contact_phone")),
                _to_text(row.get("assignee")) or "미배정",
                _format_datetime(row.get("scheduled_visit_at")),
                _bool_label(row.get("recurrence_flag")),
                int(row.get("recurrence_count") or 0),
                _to_text(row.get("source_workbook")),
                _to_text(row.get("source_sheet")),
                _to_text(row.get("source_row_number")),
                _to_text(row.get("import_batch_id")),
            ]
        )
    return rows


def _add_summary_sheet(wb: Workbook, *, site: str, cases: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "요약"
    status_counts = Counter(_to_text(row.get("status")) for row in cases)
    type_counts = Counter(_to_text(row.get("complaint_type")) for row in cases)
    unique_households = {(_to_text(row.get("building")), _to_text(row.get("unit_number"))) for row in cases}
    recurrence_cases = [
        row
        for row in cases
        if bool(row.get("recurrence_flag")) or int(row.get("recurrence_count") or 0) > 0 or _to_text(row.get("status")) == "reopened"
    ]
    open_cases = [row for row in cases if _to_text(row.get("status")) in ACTIVE_STATUSES]
    building_counts = Counter(_to_text(row.get("building")) for row in cases)

    summary_rows = [
        ["기준단지", site],
        ["생성일시", datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")],
        ["전체 민원", len(cases)],
        ["미처리/진행중", len(open_cases)],
        ["재민원 점검대상", len(recurrence_cases)],
        ["민원 발생 세대수", len(unique_households)],
        [],
        ["상태", "건수"],
    ]
    for status, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0])):
        summary_rows.append([STATUS_LABELS.get(status, status), count])

    summary_rows.extend([[], ["유형", "건수"]])
    for complaint_type, count in sorted(type_counts.items(), key=lambda item: (-item[1], item[0])):
        summary_rows.append([complaint_type_label(complaint_type), count])

    summary_rows.extend([[], ["동", "건수"]])
    for building, count in sorted(building_counts.items(), key=lambda item: _sort_key(item[0], "")):
        summary_rows.append([building, count])

    for row in summary_rows:
        ws.append(row)

    for cell in ws["A1:B1"]:
        pass
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18


def export_report(*, site: str, output_path: Path) -> Path:
    cases = _load_cases(site)
    wb = Workbook()
    _add_summary_sheet(wb, site=site, cases=cases)

    household_headers = [
        "동",
        "호수",
        "총민원수",
        "미처리건수",
        "재민원표시건수",
        "최근상태",
        "최초접수일시",
        "최근접수일시",
        "최근민원유형",
        "최근제목",
        "담당자",
        "연락처",
        "유형묶음",
        "원본시트",
        "원본행",
    ]
    open_headers = [
        "민원ID",
        "상태",
        "우선순위",
        "접수일시",
        "동",
        "호수",
        "민원유형",
        "제목",
        "상세내용",
        "입주민명",
        "연락처",
        "담당자",
        "방문예정일시",
        "재민원여부",
        "재민원횟수",
        "원본파일",
        "원본시트",
        "원본행",
        "가져오기배치",
    ]

    household_rows = _build_household_rows(cases)
    open_cases = [row for row in cases if _to_text(row.get("status")) in ACTIVE_STATUSES]
    recurrence_cases = [
        row
        for row in cases
        if bool(row.get("recurrence_flag")) or int(row.get("recurrence_count") or 0) > 0 or _to_text(row.get("status")) == "reopened"
    ]

    _append_sheet(wb.create_sheet("동호수현황"), household_headers, household_rows)
    _append_sheet(wb.create_sheet("미처리점검"), open_headers, _build_case_rows(open_cases))
    _append_sheet(wb.create_sheet("재민원점검"), open_headers, _build_case_rows(recurrence_cases))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    args = _parse_args()
    output_path = Path(args.output).expanduser().resolve()
    exported = export_report(site=args.site, output_path=output_path)
    print(exported)


if __name__ == "__main__":
    main()
