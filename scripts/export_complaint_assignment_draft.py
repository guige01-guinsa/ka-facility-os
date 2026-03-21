from __future__ import annotations

import argparse
import math
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy import select

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.database import complaint_cases, get_conn
from app.domains.complaints.service import PRIORITY_LABELS, STATUS_LABELS, complaint_type_label


KST = timezone(timedelta(hours=9))
ACTIVE_STATUSES = {"received", "assigned", "visit_scheduled", "in_progress", "reopened"}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EDIT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export initial complaint assignment draft workbook")
    parser.add_argument("--site", required=True, help="site name, for example 연산더샵")
    parser.add_argument("--output", required=True, help="xlsx output path")
    parser.add_argument("--crew-count", type=int, default=4, help="number of field crews for the draft assignment")
    return parser.parse_args()


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_datetime(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.strftime("%Y-%m-%d %H:%M")
        return value.astimezone(KST).strftime("%Y-%m-%d %H:%M")
    return _to_text(value)


def _building_num(building: Any) -> int | None:
    digits = "".join(ch for ch in _to_text(building) if ch.isdigit())
    return int(digits) if digits else None


def _sort_key(building: Any, unit_number: Any) -> tuple[Any, ...]:
    building_num = _building_num(building)
    unit_digits = "".join(ch for ch in _to_text(unit_number) if ch.isdigit())
    unit_num = int(unit_digits) if unit_digits else math.inf
    return (
        building_num if building_num is not None else math.inf,
        _to_text(building),
        unit_num,
        _to_text(unit_number),
    )


def _autosize(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(text) + 2, 44))
        ws.column_dimensions[letter].width = width


def _style_header(ws: Any) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _load_active_cases(site: str) -> list[dict[str, Any]]:
    with get_conn() as conn:
        rows = conn.execute(
            select(complaint_cases)
            .where(complaint_cases.c.site == site)
            .order_by(complaint_cases.c.building.asc(), complaint_cases.c.unit_number.asc(), complaint_cases.c.reported_at.asc())
        ).mappings()
        return [dict(row) for row in rows if _to_text(row.get("status")) in ACTIVE_STATUSES]


def _partition_contiguous(building_counts: list[tuple[str, int]], crew_count: int) -> list[list[tuple[str, int]]]:
    if not building_counts:
        return []
    crew_count = max(1, min(int(crew_count), len(building_counts)))
    values = [count for _, count in building_counts]
    total = sum(values)
    target = total / crew_count
    prefix = [0]
    for value in values:
        prefix.append(prefix[-1] + value)

    inf = float("inf")
    dp = [[inf] * (crew_count + 1) for _ in range(len(values) + 1)]
    cut = [[-1] * (crew_count + 1) for _ in range(len(values) + 1)]
    dp[0][0] = 0.0
    for i in range(1, len(values) + 1):
        for k in range(1, crew_count + 1):
            for p in range(k - 1, i):
                group_sum = prefix[i] - prefix[p]
                cost = dp[p][k - 1] + (group_sum - target) ** 2
                if cost < dp[i][k]:
                    dp[i][k] = cost
                    cut[i][k] = p

    groups: list[list[tuple[str, int]]] = []
    i = len(values)
    k = crew_count
    while k > 0:
        p = cut[i][k]
        if p < 0:
            break
        groups.append(building_counts[p:i])
        i = p
        k -= 1
    groups.reverse()
    return groups


def _build_assignment_plan(cases: list[dict[str, Any]], crew_count: int) -> dict[str, str]:
    building_counter = Counter(_to_text(row.get("building")) for row in cases)
    numeric_buildings: list[tuple[str, int]] = []
    special_buildings: list[str] = []
    for building, count in sorted(building_counter.items(), key=lambda item: (_building_num(item[0]) if _building_num(item[0]) is not None else math.inf, item[0])):
        if _building_num(building) is None:
            special_buildings.append(building)
        else:
            numeric_buildings.append((building, count))

    plan: dict[str, str] = {}
    for idx, group in enumerate(_partition_contiguous(numeric_buildings, crew_count), start=1):
        for building, _count in group:
            plan[building] = f"현장{idx}"
    for building in special_buildings:
        plan[building] = "관리검토"
    return plan


def _build_instruction_sheet(wb: Workbook, *, site: str, crew_count: int, case_count: int) -> None:
    ws = wb.active
    ws.title = "배정안내"
    rows = [
        ["항목", "내용"],
        ["대상 단지", site],
        ["생성일시", datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")],
        ["배정 기준", f"{crew_count}개 현장조 기준 동선 최소화 + 건수 균형"],
        ["대상 민원 수", case_count],
        ["사용 방법 1", "업데이트양식 시트의 추천담당자와 변경상태코드를 검토합니다."],
        ["사용 방법 2", "적용할 행은 적용여부(Y)에 Y를 입력합니다."],
        ["사용 방법 3", "추천안 그대로 반영하려면 변경담당자/변경상태코드를 유지하면 됩니다."],
        ["사용 방법 4", "방문 예정만 추가할 경우 변경방문예정만 입력하면 됩니다."],
        ["주의", "동 값이 비정상인 행은 관리검토로 분리했습니다."],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT
    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_summary_sheet(wb: Workbook, *, cases: list[dict[str, Any]], plan: dict[str, str]) -> None:
    ws = wb.create_sheet("배정요약")
    ws.append(["담당자", "배정동", "민원건수", "세대수", "주요유형", "비고"])
    grouped_cases: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cases:
        grouped_cases[plan.get(_to_text(row.get("building")), "관리검토")].append(row)

    building_by_assignee: dict[str, list[str]] = defaultdict(list)
    for building, assignee in plan.items():
        building_by_assignee[assignee].append(building)

    assignees = sorted(grouped_cases.keys(), key=lambda name: (name == "관리검토", name))
    for assignee in assignees:
        rows = grouped_cases[assignee]
        households = {(_to_text(row.get("building")), _to_text(row.get("unit_number"))) for row in rows}
        type_counts = Counter(complaint_type_label(_to_text(row.get("complaint_type"))) for row in rows)
        top_types = ", ".join(f"{label} {count}" for label, count in type_counts.most_common(3))
        buildings = ", ".join(sorted(building_by_assignee.get(assignee, []), key=lambda item: (_building_num(item) if _building_num(item) is not None else math.inf, item)))
        note = "원본 동 값 확인 필요" if assignee == "관리검토" else ""
        ws.append([assignee, buildings, len(rows), len(households), top_types, note])

    _style_header(ws)


def _build_update_sheet(wb: Workbook, *, cases: list[dict[str, Any]], plan: dict[str, str]) -> None:
    ws = wb.create_sheet("업데이트양식")
    headers = [
        "민원ID",
        "단지",
        "동",
        "호수",
        "현재상태코드",
        "현재상태",
        "민원유형",
        "제목",
        "상세내용",
        "입주민명",
        "연락처",
        "현재담당자",
        "현재방문예정",
        "현재재민원표시",
        "적용여부(Y)",
        "변경상태코드",
        "변경담당자",
        "변경방문예정",
        "재민원표시",
        "처리메모",
        "작업자비고",
        "추천담당자",
        "추천근거",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    ordered = sorted(cases, key=lambda row: (_sort_key(row.get("building"), row.get("unit_number")), row.get("reported_at") or row.get("created_at") or datetime.min.replace(tzinfo=timezone.utc), row.get("id") or 0))
    for row in ordered:
        status_code = _to_text(row.get("status"))
        suggested_assignee = plan.get(_to_text(row.get("building")), "관리검토")
        suggested_status = "assigned" if status_code == "received" else ""
        rationale = "동 기준 균형 배정" if suggested_assignee != "관리검토" else "원본 동 값 누락/비정상 확인 필요"
        ws.append(
            [
                row.get("id"),
                _to_text(row.get("site")),
                _to_text(row.get("building")),
                _to_text(row.get("unit_number")),
                status_code,
                STATUS_LABELS.get(status_code, status_code),
                complaint_type_label(_to_text(row.get("complaint_type"))),
                _to_text(row.get("title")),
                _to_text(row.get("description")),
                _to_text(row.get("resident_name")),
                _to_text(row.get("contact_phone")),
                _to_text(row.get("assignee")),
                _format_datetime(row.get("scheduled_visit_at")),
                "예" if bool(row.get("recurrence_flag")) else "아니오",
                "",
                suggested_status,
                suggested_assignee,
                "",
                "",
                "",
                "",
                suggested_assignee,
                rationale,
            ]
        )

    for row_idx in range(2, ws.max_row + 1):
        for col in ["O", "P", "Q", "R", "S", "T", "U"]:
            ws[f"{col}{row_idx}"].fill = EDIT_FILL

    apply_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    status_validation = DataValidation(
        type="list",
        formula1='"received,assigned,visit_scheduled,in_progress,resolved,resident_confirmed,reopened,closed"',
        allow_blank=True,
    )
    recurrence_validation = DataValidation(type="list", formula1='"예,아니오"', allow_blank=True)
    ws.add_data_validation(apply_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(recurrence_validation)
    apply_validation.add(f"O2:O{ws.max_row}")
    status_validation.add(f"P2:P{ws.max_row}")
    recurrence_validation.add(f"S2:S{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def export_assignment_draft(*, site: str, output_path: Path, crew_count: int) -> dict[str, Any]:
    cases = _load_active_cases(site)
    plan = _build_assignment_plan(cases, crew_count)
    wb = Workbook()
    _build_instruction_sheet(wb, site=site, crew_count=crew_count, case_count=len(cases))
    _build_summary_sheet(wb, cases=cases, plan=plan)
    _build_update_sheet(wb, cases=cases, plan=plan)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    summary_rows: dict[str, dict[str, Any]] = {}
    households_by_assignee: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for row in cases:
        assignee = plan.get(_to_text(row.get("building")), "관리검토")
        info = summary_rows.setdefault(assignee, {"case_count": 0, "buildings": set()})
        info["case_count"] += 1
        info["buildings"].add(_to_text(row.get("building")))
        households_by_assignee[assignee].add((_to_text(row.get("building")), _to_text(row.get("unit_number"))))

    normalized_summary = {
        assignee: {
            "case_count": values["case_count"],
            "household_count": len(households_by_assignee[assignee]),
            "buildings": sorted(values["buildings"], key=lambda item: (_building_num(item) if _building_num(item) is not None else math.inf, item)),
        }
        for assignee, values in summary_rows.items()
    }
    return {"path": str(output_path), "summary": normalized_summary}


def main() -> None:
    args = _parse_args()
    output_path = Path(args.output).expanduser().resolve()
    result = export_assignment_draft(site=args.site, output_path=output_path, crew_count=int(args.crew_count))
    print(result["path"])
    for assignee, payload in sorted(result["summary"].items(), key=lambda item: (item[0] == "관리검토", item[0])):
        print(f"{assignee}: cases={payload['case_count']}, households={payload['household_count']}, buildings={', '.join(payload['buildings'])}")


if __name__ == "__main__":
    main()
