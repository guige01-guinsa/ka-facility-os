from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy import select

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.database import complaint_cases, get_conn
from app.domains.complaints.service import STATUS_LABELS, complaint_type_label


KST = timezone(timedelta(hours=9))
ACTIVE_STATUSES = {"received", "assigned", "visit_scheduled", "in_progress", "reopened"}
EDITABLE_COLUMNS = {"O", "P", "Q", "R", "S", "T", "U"}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export complaint bulk update workbook")
    parser.add_argument("--site", required=True, help="site name, for example 연산더샵")
    parser.add_argument("--output", required=True, help="xlsx output path")
    parser.add_argument(
        "--include-closed",
        action="store_true",
        help="include closed/resolved complaints too; default exports only active complaints",
    )
    return parser.parse_args()


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_datetime(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.strftime("%Y-%m-%d %H:%M")
        return value.astimezone(KST).strftime("%Y-%m-%d %H:%M")
    return _to_text(value)


def _sort_key(building: Any, unit_number: Any) -> tuple[Any, ...]:
    def _token(value: Any) -> tuple[int, str]:
        text = _to_text(value)
        digits = "".join(ch for ch in text if ch.isdigit())
        return (int(digits), text) if digits else (10**9, text)

    return (*_token(building), *_token(unit_number))


def _autosize(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(text) + 2, 44))
        ws.column_dimensions[letter].width = width


def _load_cases(site: str, *, include_closed: bool) -> list[dict[str, Any]]:
    with get_conn() as conn:
        stmt = (
            select(complaint_cases)
            .where(complaint_cases.c.site == site)
            .order_by(complaint_cases.c.building.asc(), complaint_cases.c.unit_number.asc(), complaint_cases.c.reported_at.asc())
        )
        rows = [dict(row) for row in conn.execute(stmt).mappings()]
    if include_closed:
        return sorted(rows, key=lambda row: (_sort_key(row.get("building"), row.get("unit_number")), row.get("id") or 0))
    return [
        row
        for row in sorted(rows, key=lambda row: (_sort_key(row.get("building"), row.get("unit_number")), row.get("id") or 0))
        if _to_text(row.get("status")) in ACTIVE_STATUSES
    ]


def _build_instruction_sheet(wb: Workbook, *, site: str, count: int) -> None:
    ws = wb.active
    ws.title = "작업안내"
    rows = [
        ["항목", "내용"],
        ["대상 단지", site],
        ["양식 생성일시", datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")],
        ["대상 민원 수", count],
        ["사용 순서 1", "업데이트양식 시트에서 노란색 칸만 수정합니다."],
        ["사용 순서 2", "적용할 행은 '적용여부(Y)' 칸에 Y를 입력합니다."],
        ["사용 순서 3", "변경상태코드에는 received, assigned, visit_scheduled, in_progress, resolved, resident_confirmed, reopened, closed 중 하나를 입력합니다."],
        ["사용 순서 4", "변경담당자는 비우면 변경 없음이고, '삭제'를 입력하면 담당자를 비웁니다."],
        ["사용 순서 5", "변경방문예정은 '2026-03-22 14:00' 형식으로 입력하고, '삭제'를 입력하면 방문예정을 비웁니다."],
        ["사용 순서 6", "재민원표시는 예/아니오 중 하나를 입력합니다. 비우면 변경하지 않습니다."],
        ["사용 순서 7", "처리메모를 적으면 민원 이력에 별도 메모 이벤트로 남습니다."],
        ["적용 명령 예시", ".\\.venv\\Scripts\\python.exe scripts\\apply_complaint_bulk_update_template.py --workbook \"c:\\연산더샵\\시설과장\\민원처리_일괄업데이트_20260322.xlsx\""],
        ["실반영 명령 예시", ".\\.venv\\Scripts\\python.exe scripts\\apply_complaint_bulk_update_template.py --workbook \"c:\\연산더샵\\시설과장\\민원처리_일괄업데이트_20260322.xlsx\" --apply --actor 현장1"],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_code_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("코드안내")
    rows = [["상태코드", "상태명"]]
    for code, label in STATUS_LABELS.items():
        rows.append([code, label])
    rows.extend(
        [
            [],
            ["적용여부", "Y 또는 N"],
            ["재민원표시", "예 또는 아니오"],
            ["삭제표시", "삭제 또는 CLEAR"],
        ]
    )
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    _autosize(ws)


def _build_update_sheet(wb: Workbook, cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("업데이트양식")
    headers = [
        "민원ID",
        "단지",
        "동",
        "호수",
        "현재상태코드",
        "현재상태",
        "민원유형",
        "제목",
        "상세내용",
        "입주민명",
        "연락처",
        "현재담당자",
        "현재방문예정",
        "현재재민원표시",
        "적용여부(Y)",
        "변경상태코드",
        "변경담당자",
        "변경방문예정",
        "재민원표시",
        "처리메모",
        "작업자비고",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in cases:
        status_code = _to_text(row.get("status"))
        ws.append(
            [
                row.get("id"),
                _to_text(row.get("site")),
                _to_text(row.get("building")),
                _to_text(row.get("unit_number")),
                status_code,
                STATUS_LABELS.get(status_code, status_code),
                complaint_type_label(_to_text(row.get("complaint_type"))),
                _to_text(row.get("title")),
                _to_text(row.get("description")),
                _to_text(row.get("resident_name")),
                _to_text(row.get("contact_phone")),
                _to_text(row.get("assignee")),
                _format_datetime(row.get("scheduled_visit_at")),
                "예" if bool(row.get("recurrence_flag")) else "아니오",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    for row_idx in range(2, ws.max_row + 1):
        for col in EDITABLE_COLUMNS:
            ws[f"{col}{row_idx}"].fill = editable_fill

    apply_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    status_validation = DataValidation(
        type="list",
        formula1='"received,assigned,visit_scheduled,in_progress,resolved,resident_confirmed,reopened,closed"',
        allow_blank=True,
    )
    recurrence_validation = DataValidation(type="list", formula1='"예,아니오"', allow_blank=True)
    ws.add_data_validation(apply_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(recurrence_validation)
    apply_validation.add(f"O2:O{ws.max_row}")
    status_validation.add(f"P2:P{ws.max_row}")
    recurrence_validation.add(f"S2:S{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def export_template(*, site: str, output_path: Path, include_closed: bool) -> Path:
    cases = _load_cases(site, include_closed=include_closed)
    wb = Workbook()
    _build_instruction_sheet(wb, site=site, count=len(cases))
    _build_code_sheet(wb)
    _build_update_sheet(wb, cases)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    args = _parse_args()
    output_path = Path(args.output).expanduser().resolve()
    exported = export_template(site=args.site, output_path=output_path, include_closed=bool(args.include_closed))
    print(exported)


if __name__ == "__main__":
    main()
