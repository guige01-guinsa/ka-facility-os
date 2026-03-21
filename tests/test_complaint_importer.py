from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


def _write_headers(sheet, *, title: str) -> None:
    sheet["A1"] = title
    sheet["A2"] = "순번"
    sheet["B2"] = "날짜"
    sheet["C2"] = "동"
    sheet["D2"] = "호수"
    sheet["E2"] = "민원내용"
    sheet["F2"] = "전화번호"


def test_complaint_importer_reads_multi_sheet_and_dedupes(tmp_path: Path) -> None:
    from app.domains.complaints.importer import dedupe_rows, iter_workbook_rows

    workbook = Workbook()
    ws_total = workbook.active
    ws_total.title = "전체"
    _write_headers(ws_total, title="전체 민원")
    ws_total.append([1, "2025-07-22", 101, 503, "방충망 페인트오염", "010-1111-2222"])

    ws_101 = workbook.create_sheet("101동")
    _write_headers(ws_101, title="101동 민원")
    ws_101.append([1, "2025-07-22", 101, 503, "방충망 페인트오염", "010-1111-2222"])
    ws_101.append([2, "2025-11-12", 101, 402, "안방방충망오염, 거실 난간오염", "010-8903-0055"])

    ws_louver = workbook.create_sheet("루버창")
    ws_louver.append([None, None, None, None, None, None])
    ws_louver.append([None, None, None, None, None, None])
    ws_louver.append([None, "2025-11-17", 107, 804, "페인트작업중 실외기실 루버창 찌그러짐", "010-2721-5405"])

    ws_glass = workbook.create_sheet("유리")
    _write_headers(ws_glass, title="유리 민원")
    ws_glass.append([1, None, 109, 1801, "안방베란다 유리 교체요", "010-8459-0808"])
    ws_glass.append([2, "추가접수", None, None, None, None])

    path = tmp_path / "complaints.xlsx"
    workbook.save(path)

    rows = iter_workbook_rows(path, site="연산더샵")
    unique_rows, duplicates = dedupe_rows(rows)

    assert len(rows) == 5
    assert len(unique_rows) == 4
    assert sum(len(items) for items in duplicates.values()) == 1
    assert any(row.building == "107동" and row.unit_number == "804호" for row in unique_rows)
    assert any(row.complaint_type == "louver_issue" for row in unique_rows)
