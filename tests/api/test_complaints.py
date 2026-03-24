from __future__ import annotations

from io import BytesIO
from datetime import datetime, timezone

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4

from app.domains.complaints import reporting
from app.web.complaints import complaints_script_version
from tests.helpers.common import _owner_headers


def test_complaints_mobile_page_renders_field_console(app_client: TestClient) -> None:
    page = app_client.get("/web/complaints")
    assert page.status_code == 200
    assert page.headers["content-type"].startswith("text/html")
    assert page.headers.get("cache-control") == "no-store"
    assert page.headers.get("pragma") == "no-cache"
    assert page.headers.get("x-robots-tag") == "noindex, nofollow"
    assert "Field Workflow Console" in page.text
    assert "нҳ„мһҘ нҒҗ" in page.text
    assert "лҜјмӣҗ мӢ к·ң л“ұлЎқ" in page.text
    assert "м—°кІ° нҷ•мқё" in page.text
    assert "нҶ нҒ° ліҙкё°" in page.text
    assert "м—‘м…Җ м¶ңл Ҙ" in page.text
    assert "PDF м¶ңл Ҙ" in page.text
    assert "лҸҷ/нҳё мҲң" in page.text
    assert "л¶„лҘҳ/лҸҷ/нҳё мҲң" in page.text
    assert "к·ёлЈ№ н‘ңмӢң" in page.text
    assert "DB л ҲмҪ”л“ң кҙҖлҰ¬" in page.text
    assert "м№јлҹј мҲЁк№Җ/н‘ңмӢң" in page.text
    assert "м „мІҙ н‘ңмӢң" in page.text
    assert "кё°ліёк°’ ліөмӣҗ" in page.text
    assert "м¶ңл Ҙ н‘ңм§Җ м„Өм •" in page.text
    assert "нҡҢмӮ¬лӘ… м„ нғқ" in page.text
    assert "кіөмӮ¬м—…мІҙ м„ нғқ" in page.text
    assert "лЎңкі  мқҙлҜём§Җ л¶Ҳлҹ¬мҳӨкё°" in page.text
    assert "н‘ңм§Җ н”„лҰ¬м…Ӣ" in page.text
    assert "кҙҖлҰ¬мһҗ кіөнҶө кё°ліёк°’" in page.text
    assert "н‘ңм§Җ лҜёлҰ¬ліҙкё°" in page.text
    assert f'/web/complaints/app.js?v={complaints_script_version()}' in page.text


def test_complaints_mobile_script_is_cacheable(app_client: TestClient) -> None:
    asset = app_client.get(f"/web/complaints/app.js?v={complaints_script_version()}")
    assert asset.status_code == 200
    assert asset.headers["content-type"].startswith("application/javascript")
    assert asset.headers.get("cache-control") == "public, max-age=31536000, immutable"
    assert asset.headers.get("etag") == complaints_script_version()
    assert asset.headers.get("x-content-type-options") == "nosniff"
    assert "kaFacility.auth.token" in asset.text
    assert "kaFacility.auth.profile" in asset.text
    assert "м„ёлҢҖ мқҙл Ҙ" in asset.text
    assert "л¬ёмһҗ л°ңмҶЎ" in asset.text
    assert "л№„мҡ© мһ…л Ҙ" in asset.text


def test_complaint_case_crud_and_household_history(app_client: TestClient) -> None:
    headers = _owner_headers()
    created = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "101",
            "unit_number": "503",
            "description": "л°©м¶©л§қ нҺҳмқёнҠёмҳӨм—ј л°Ҹ лӮңк°„ мҳӨм—ј",
            "priority": "high",
            "contact_phone": "01058961551",
        },
    )
    assert created.status_code == 201
    body = created.json()
    complaint_id = body["id"]
    assert body["building"] == "101лҸҷ"
    assert body["unit_number"] == "503нҳё"
    assert body["complaint_type"] == "composite"
    assert body["priority"] == "high"
    assert body["status"] == "received"

    listed = app_client.get(
        "/api/complaints?site=м—°мӮ°лҚ”мғө&building=101лҸҷ&unit_number=503нҳё",
        headers=headers,
    )
    assert listed.status_code == 200
    rows = listed.json()
    assert len(rows) == 1
    assert rows[0]["id"] == complaint_id

    visit_at = datetime(2026, 3, 21, 5, 0, tzinfo=timezone.utc).isoformat()
    updated = app_client.patch(
        f"/api/complaints/{complaint_id}",
        headers=headers,
        json={
            "status": "visit_scheduled",
            "assignee": "нҳ„мһҘл°ҳмһҘ",
            "scheduled_visit_at": visit_at,
        },
    )
    assert updated.status_code == 200
    updated_body = updated.json()
    assert updated_body["status"] == "visit_scheduled"
    assert updated_body["assignee"] == "нҳ„мһҘл°ҳмһҘ"

    event = app_client.post(
        f"/api/complaints/{complaint_id}/events",
        headers=headers,
        json={
            "event_type": "field_visit",
            "to_status": "resolved",
            "note": "л°©л¬ё нӣ„ лӮңк°„кіј л°©м¶©л§қ мҳӨм—ј м ңкұ°",
            "detail": {"worker": "2мқё1мЎ°"},
        },
    )
    assert event.status_code == 201
    event_body = event.json()
    assert event_body["event_type"] == "field_visit"
    assert event_body["to_status"] == "resolved"

    updated_event = app_client.patch(
        f"/api/complaints/events/{event_body['id']}",
        headers=headers,
        json={
            "event_type": "rework",
            "note": "мһ¬л°©л¬ё мӨҖл№„",
            "detail": {"worker": "1мқё1мЎ°"},
        },
    )
    assert updated_event.status_code == 200
    assert updated_event.json()["event_type"] == "rework"
    assert updated_event.json()["note"] == "мһ¬л°©л¬ё мӨҖл№„"

    detail = app_client.get(f"/api/complaints/{complaint_id}", headers=headers)
    assert detail.status_code == 200
    detail_body = detail.json()
    assert detail_body["case"]["status"] == "resolved"
    assert len(detail_body["events"]) >= 3

    history = app_client.get(
        "/api/complaints/households/history?site=м—°мӮ°лҚ”мғө&building=101лҸҷ&unit_number=503нҳё",
        headers=headers,
    )
    assert history.status_code == 200
    history_body = history.json()
    assert history_body["site"] == "м—°мӮ°лҚ”мғө"
    assert len(history_body["complaints"]) == 1
    assert history_body["complaints"][0]["id"] == complaint_id

    deleted_event = app_client.delete(
        f"/api/complaints/events/{event_body['id']}",
        headers=headers,
    )
    assert deleted_event.status_code == 200
    assert deleted_event.json()["deleted"] is True

    deleted_case = app_client.delete(
        f"/api/complaints/{complaint_id}",
        headers=headers,
    )
    assert deleted_case.status_code == 200
    assert deleted_case.json()["deleted"] is True

    after_delete = app_client.get(
        "/api/complaints?site=м—°мӮ°лҚ”мғө&building=101лҸҷ&unit_number=503нҳё",
        headers=headers,
    )
    assert after_delete.status_code == 200
    assert after_delete.json() == []

    deleted_detail = app_client.get(f"/api/complaints/{complaint_id}", headers=headers)
    assert deleted_detail.status_code == 404


def test_complaint_attachment_message_and_cost_workflow(app_client: TestClient) -> None:
    headers = _owner_headers()
    created = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "102лҸҷ",
            "unit_number": "1204нҳё",
            "description": "кұ°мӢӨ л°©м¶©л§қ мҳӨм—ј",
            "contact_phone": "010-8529-4439",
        },
    )
    assert created.status_code == 201
    complaint_id = created.json()["id"]

    blocked_upload = app_client.post(
        f"/api/complaints/{complaint_id}/attachments",
        headers=headers,
        data={"attachment_kind": "intake", "note": "html not allowed"},
        files={"file": ("bad.html", b"<script>alert(1)</script>", "text/html")},
    )
    assert blocked_upload.status_code == 415

    uploaded = app_client.post(
        f"/api/complaints/{complaint_id}/attachments",
        headers=headers,
        data={"attachment_kind": "intake", "note": "м ‘мҲҳ мӮ¬м§„"},
        files={"file": ("complaint-photo.txt", b"complaint evidence", "text/plain")},
    )
    assert uploaded.status_code == 201
    attachment = uploaded.json()
    attachment_id = attachment["id"]
    assert attachment["attachment_kind"] == "intake"
    assert attachment["file_name"] == "complaint-photo.txt"

    attachments = app_client.get(
        f"/api/complaints/{complaint_id}/attachments",
        headers=headers,
    )
    assert attachments.status_code == 200
    attachment_rows = attachments.json()
    assert len(attachment_rows) == 1
    assert attachment_rows[0]["id"] == attachment_id

    downloaded = app_client.get(
        f"/api/complaints/attachments/{attachment_id}/download",
        headers=headers,
    )
    assert downloaded.status_code == 200
    assert downloaded.content == b"complaint evidence"
    assert len(downloaded.headers["x-complaint-sha256"]) == 64

    sent = app_client.post(
        f"/api/complaints/{complaint_id}/messages",
        headers=headers,
        json={"body": "лҜјмӣҗ м ‘мҲҳлҗҳм—ҲмҠөлӢҲлӢӨ. л°©л¬ё мқјм •мқ„ лӢӨмӢң м•ҲлӮҙл“ңлҰ¬кІ мҠөлӢҲлӢӨ."},
    )
    assert sent.status_code == 201
    message = sent.json()
    assert message["delivery_status"] == "sent"
    assert message["recipient"] == "010-8529-4439"

    updated_attachment = app_client.patch(
        f"/api/complaints/attachments/{attachment_id}",
        headers=headers,
        json={"attachment_kind": "after", "note": "мһ‘м—… нӣ„ мӮ¬м§„"},
    )
    assert updated_attachment.status_code == 200
    assert updated_attachment.json()["attachment_kind"] == "after"
    assert updated_attachment.json()["note"] == "мһ‘м—… нӣ„ мӮ¬м§„"

    updated_message = app_client.patch(
        f"/api/complaints/messages/{message['id']}",
        headers=headers,
        json={
            "recipient": "01077778888",
            "template_key": "visit_notice",
            "delivery_status": "sent",
            "body": "лӮҙмқј л°©л¬ё мҳҲм •мһ…лӢҲлӢӨ.",
        },
    )
    assert updated_message.status_code == 200
    assert updated_message.json()["recipient"] == "010-7777-8888"
    assert updated_message.json()["template_key"] == "visit_notice"
    assert updated_message.json()["body"] == "лӮҙмқј л°©л¬ё мҳҲм •мһ…лӢҲлӢӨ."

    cost = app_client.post(
        f"/api/complaints/{complaint_id}/cost-items",
        headers=headers,
        json={
            "cost_category": "cleaning",
            "item_name": "л°©м¶©л§қ мІӯмҶҢ",
            "quantity": 2,
            "unit_price": 10000,
            "material_cost": 3000,
            "labor_cost": 2000,
            "vendor_cost": 0,
        },
    )
    assert cost.status_code == 201
    cost_item = cost.json()
    assert cost_item["total_cost"] == 25000.0

    updated_cost = app_client.patch(
        f"/api/complaints/cost-items/{cost_item['id']}",
        headers=headers,
        json={
            "quantity": 3,
            "unit_price": 9000,
            "note": "3л©ҙ мһ‘м—…мңјлЎң мҲҳм •",
        },
    )
    assert updated_cost.status_code == 200
    assert updated_cost.json()["quantity"] == 3.0
    assert updated_cost.json()["total_cost"] == 32000.0
    assert updated_cost.json()["note"] == "3л©ҙ мһ‘м—…мңјлЎң мҲҳм •"

    detail = app_client.get(f"/api/complaints/{complaint_id}", headers=headers)
    assert detail.status_code == 200
    detail_body = detail.json()
    assert len(detail_body["attachments"]) == 1
    assert len(detail_body["messages"]) == 1
    assert len(detail_body["cost_items"]) == 1
    assert detail_body["total_cost"] == 32000.0

    deleted_attachment = app_client.delete(
        f"/api/complaints/attachments/{attachment_id}",
        headers=headers,
    )
    assert deleted_attachment.status_code == 200
    assert deleted_attachment.json()["deleted"] is True

    deleted_message = app_client.delete(
        f"/api/complaints/messages/{message['id']}",
        headers=headers,
    )
    assert deleted_message.status_code == 200
    assert deleted_message.json()["deleted"] is True

    deleted_cost = app_client.delete(
        f"/api/complaints/cost-items/{cost_item['id']}",
        headers=headers,
    )
    assert deleted_cost.status_code == 200
    assert deleted_cost.json()["deleted"] is True

    after_delete_detail = app_client.get(f"/api/complaints/{complaint_id}", headers=headers)
    assert after_delete_detail.status_code == 200
    after_delete_body = after_delete_detail.json()
    assert after_delete_body["attachments"] == []
    assert after_delete_body["messages"] == []
    assert after_delete_body["cost_items"] == []
    assert after_delete_body["total_cost"] == 0.0


def test_complaint_admin_record_grid_uses_server_side_limit_and_count(app_client: TestClient) -> None:
    headers = _owner_headers()
    token = "limit-search-token-20260324"
    first = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "106лҸҷ",
            "unit_number": "601нҳё",
            "description": token + " мІ« лІҲм§ё",
        },
    )
    assert first.status_code == 201

    second = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "106лҸҷ",
            "unit_number": "602нҳё",
            "description": token + " л‘җ лІҲм§ё",
        },
    )
    assert second.status_code == 201

    listed = app_client.get(
        f"/api/complaints/admin/records?site=м—°мӮ°лҚ”мғө&record_type=cases&limit=1&q={token}",
        headers=headers,
    )
    assert listed.status_code == 200
    body = listed.json()
    assert body["total_count"] == 2
    assert len(body["rows"]) == 1
    assert body["rows"][0]["id"] == second.json()["id"]


def test_complaint_report_exports(app_client: TestClient) -> None:
    headers = _owner_headers()
    first = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "103лҸҷ",
            "unit_number": "901нҳё",
            "description": "кұ°мӢӨ л°©м¶©л§қ мҳӨм—ј",
            "contact_phone": "010-1111-2222",
        },
    )
    assert first.status_code == 201

    second = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "104лҸҷ",
            "unit_number": "1201нҳё",
            "description": "лӮңк°„ мҳӨм—ј",
            "contact_phone": "010-3333-4444",
        },
    )
    assert second.status_code == 201
    complaint_id = second.json()["id"]

    resolved = app_client.patch(
        f"/api/complaints/{complaint_id}",
        headers=headers,
        json={"status": "closed", "assignee": "нҳ„мһҘ3"},
    )
    assert resolved.status_code == 200

    xlsx_resp = app_client.get(
        "/api/complaints/reports/xlsx?site=м—°мӮ°лҚ”мғө&report_type=unresolved&sort_by=building_unit",
        headers=headers,
    )
    assert xlsx_resp.status_code == 200
    assert xlsx_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(xlsx_resp.content))
    assert workbook.sheetnames == [
        "мҡ”м•Ҫ",
        "лҜјмӣҗлӘ©лЎқ",
        "db_complaint_cases",
        "db_complaint_events",
        "db_complaint_attachments",
        "db_complaint_messages",
        "db_complaint_cost_items",
    ]
    summary_map = {
        workbook["мҡ”м•Ҫ"][f"A{row_idx}"].value: workbook["мҡ”м•Ҫ"][f"B{row_idx}"].value
        for row_idx in range(2, workbook["мҡ”м•Ҫ"].max_row + 1)
    }
    assert summary_map["м¶ңл Ҙкө¬л¶„"] == "лҜёмІҳлҰ¬"
    assert summary_map["м •л ¬кё°мӨҖ"] == "лҸҷ/нҳё мҲң"
    assert summary_map["к·ёлЈ№н‘ңмӢң"] == "м—ҶмқҢ"
    assert workbook["лҜјмӣҗлӘ©лЎқ"].max_row == 2
    assert workbook["db_complaint_cases"].max_row == 2
    assert workbook["db_complaint_cases"]["A2"].value == str(first.json()["id"])
    assert workbook["db_complaint_events"].max_row == 2
    assert workbook["db_complaint_events"]["B2"].value == str(first.json()["id"])
    assert workbook["db_complaint_attachments"].max_row == 1
    assert workbook["db_complaint_messages"].max_row == 1
    assert workbook["db_complaint_cost_items"].max_row == 1

    pdf_resp = app_client.get(
        "/api/complaints/reports/pdf?site=м—°мӮ°лҚ”мғө&report_type=building",
        headers=headers,
    )
    assert pdf_resp.status_code == 200
    assert pdf_resp.headers["content-type"].startswith("application/pdf")
    assert pdf_resp.content.startswith(b"%PDF")

    custom_pdf_resp = app_client.post(
        "/api/complaints/reports/pdf",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "report_type": "all",
            "building": "103лҸҷ",
            "sort_by": "building_unit",
            "group_by": "building",
            "cover": {
                "company_name": "н…ҢмҠӨнҠё мӢңм„ӨкҙҖлҰ¬",
                "contractor_name": "н…ҢмҠӨнҠё лҸ„мһҘм—…мІҙ",
                "submission_phrase": "к·Җ кҙҖлҰ¬мӮ¬л¬ҙмҶҢ кІҖнҶ лҘј мң„н•ҙ м•„лһҳмҷҖ к°ҷмқҙ м ңм¶ңн•©лӢҲлӢӨ.",
                "logo_data_url": "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+a5tQAAAAASUVORK5CYII=",
                "logo_file_name": "test-logo.png",
            },
        },
    )
    assert custom_pdf_resp.status_code == 200
    assert custom_pdf_resp.headers["content-type"].startswith("application/pdf")
    assert custom_pdf_resp.content.startswith(b"%PDF")


def test_complaint_export_report_supports_building_unit_sort(app_client: TestClient) -> None:
    headers = _owner_headers()
    higher_unit = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м •л ¬н…ҢмҠӨнҠё",
            "building": "101лҸҷ",
            "unit_number": "1202нҳё",
            "description": "лҠҰкІҢ ліҙм—¬м•ј н•ҳлҠ” м„ёлҢҖ",
        },
    )
    assert higher_unit.status_code == 201

    lower_unit = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м •л ¬н…ҢмҠӨнҠё",
            "building": "101лҸҷ",
            "unit_number": "301нҳё",
            "description": "лЁјм Җ ліҙм—¬м•ј н•ҳлҠ” м„ёлҢҖ",
        },
    )
    assert lower_unit.status_code == 201

    report = reporting.build_complaint_export_report(
        site="м •л ¬н…ҢмҠӨнҠё",
        report_type="all",
        sort_by="building_unit",
    )
    assert report.sort_by == "building_unit"
    assert report.sort_label == "лҸҷ/нҳё мҲң"
    assert report.summary_rows[3] == ("м •л ¬кё°мӨҖ", "лҸҷ/нҳё мҲң")
    assert report.summary_rows[4] == ("к·ёлЈ№н‘ңмӢң", "м—ҶмқҢ")
    assert report.rows[0][2] == "301нҳё"
    assert report.rows[1][2] == "1202нҳё"


def test_complaint_export_report_supports_category_building_unit_sort_and_grouping(app_client: TestClient) -> None:
    headers = _owner_headers()
    screen_case = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м •л ¬н…ҢмҠӨнҠё2",
            "building": "102лҸҷ",
            "unit_number": "402нҳё",
            "description": "л°©м¶©л§қ мҳӨм—ј",
        },
    )
    assert screen_case.status_code == 201

    glass_case = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м •л ¬н…ҢмҠӨнҠё2",
            "building": "101лҸҷ",
            "unit_number": "1201нҳё",
            "description": "мң лҰ¬ мҳӨм—ј",
        },
    )
    assert glass_case.status_code == 201

    report = reporting.build_complaint_export_report(
        site="м •л ¬н…ҢмҠӨнҠё2",
        report_type="all",
        sort_by="category_building_unit",
        group_by="category",
    )
    assert report.sort_by == "category_building_unit"
    assert report.sort_label == "л¶„лҘҳ/лҸҷ/нҳё мҲң"
    assert report.group_by == "category"
    assert report.group_label == "л¶„лҘҳлі„ л¬¶мқҢ"
    assert report.summary_rows[3] == ("м •л ¬кё°мӨҖ", "л¶„лҘҳ/лҸҷ/нҳё мҲң")
    assert report.summary_rows[4] == ("к·ёлЈ№н‘ңмӢң", "л¶„лҘҳлі„ л¬¶мқҢ")
    assert report.rows[0][3] == "л°©м¶©л§қ мҳӨм—ј"
    assert report.rows[1][3] == "мң лҰ¬/м°Ҫл¬ё мҳӨм—ј"


def test_complaint_pdf_cover_layout_keeps_title_below_header_block() -> None:
    layout = reporting._cover_layout(width=A4[0], height=A4[1])
    assert layout.header_bottom_y < layout.top_y
    assert layout.badge_top_y < layout.header_bottom_y
    assert layout.title_top_y < layout.badge_top_y - layout.badge_height
    assert layout.title_top_y < layout.header_bottom_y - (8 * reporting.mm)
    assert layout.badge_width == A4[0] - (2 * reporting.PDF_MARGIN)
    assert layout.badge_height > (8 * reporting.mm)


def test_complaint_pdf_table_layout_widens_contact_column() -> None:
    report = reporting.ComplaintExportReport(
        report_type="all",
        report_label="м „мІҙ",
        sort_by="reported_at",
        sort_label="м ‘мҲҳмқјмӢң мҲң",
        group_by="none",
        group_label="м—ҶмқҢ",
        site="м—°мӮ°лҚ”мғө",
        building=None,
        generated_at=datetime.now(timezone.utc),
        summary_rows=[],
        headers=["лҜјмӣҗID", "лҸҷ", "нҳёмҲҳ", "лҜјмӣҗмң нҳ•", "мғҒнғң", "лӢҙлӢ№мһҗ", "м ‘мҲҳмқјмӢң", "м—°лқҪмІҳ", "лҜјмӣҗлӮҙмҡ©"],
        rows=[],
        primary_sheet_name="лҜјмӣҗлӘ©лЎқ",
        file_stem="complaints-all-yeonsan",
        raw_sheets=[],
    )
    _, widths, _ = reporting._pdf_table_layout(report)
    assert widths[7] == 30 * reporting.mm
    assert sum(widths) <= A4[0] - (2 * reporting.PDF_MARGIN)


def test_complaint_report_cover_default_api(app_client: TestClient) -> None:
    headers = _owner_headers()
    tiny_png_data_url = (
        "data:image/png;base64,"
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+a5tQAAAAASUVORK5CYII="
    )

    global_updated = app_client.put(
        "/api/complaints/report-cover/default",
        headers=headers,
        json={
            "scope_type": "global",
            "company_name": "кёҖлЎңлІҢ мӢңм„ӨкҙҖлҰ¬",
            "contractor_name": "кёҖлЎңлІҢ лҸ„мһҘм—…мІҙ",
            "submission_phrase": "кёҖлЎңлІҢ кё°ліё м ңм¶ң л¬ёкө¬",
            "logo_data_url": tiny_png_data_url,
            "logo_file_name": "global-logo.png",
        },
    )
    assert global_updated.status_code == 200
    global_body = global_updated.json()
    assert global_body["source_scope"] == "global"
    assert global_body["company_name"] == "кёҖлЎңлІҢ мӢңм„ӨкҙҖлҰ¬"
    assert global_body["logo_present"] is True

    global_loaded = app_client.get("/api/complaints/report-cover/default", headers=headers)
    assert global_loaded.status_code == 200
    assert global_loaded.json()["company_name"] == "кёҖлЎңлІҢ мӢңм„ӨкҙҖлҰ¬"

    site_updated = app_client.put(
        "/api/complaints/report-cover/default",
        headers=headers,
        json={
            "scope_type": "site",
            "site": "м—°мӮ°лҚ”мғө",
            "company_name": "м—°мӮ°лҚ”мғө мӢңм„ӨкҙҖлҰ¬нҢҖ",
            "contractor_name": "м—°мӮ°лҚ”мғө лҸ„мһҘм—…мІҙ",
            "submission_phrase": "м—°мӮ°лҚ”мғө м ңм¶ң л¬ёкө¬",
            "clear_logo": True,
        },
    )
    assert site_updated.status_code == 200
    site_body = site_updated.json()
    assert site_body["source_scope"] == "site"
    assert site_body["site"] == "м—°мӮ°лҚ”мғө"
    assert site_body["company_name"] == "м—°мӮ°лҚ”мғө мӢңм„ӨкҙҖлҰ¬нҢҖ"
    assert site_body["logo_present"] is False

    site_loaded = app_client.get("/api/complaints/report-cover/default?site=м—°мӮ°лҚ”мғө", headers=headers)
    assert site_loaded.status_code == 200
    assert site_loaded.json()["source_scope"] == "site"
    assert site_loaded.json()["company_name"] == "м—°мӮ°лҚ”мғө мӢңм„ӨкҙҖлҰ¬нҢҖ"

    deleted_site = app_client.delete(
        "/api/complaints/report-cover/default?scope_type=site&site=м—°мӮ°лҚ”мғө",
        headers=headers,
    )
    assert deleted_site.status_code == 204

    site_fallback = app_client.get("/api/complaints/report-cover/default?site=м—°мӮ°лҚ”мғө", headers=headers)
    assert site_fallback.status_code == 200
    assert site_fallback.json()["source_scope"] == "global"
    assert site_fallback.json()["company_name"] == "кёҖлЎңлІҢ мӢңм„ӨкҙҖлҰ¬"
    assert site_fallback.json()["logo_present"] is True

    deleted_global = app_client.delete(
        "/api/complaints/report-cover/default?scope_type=global",
        headers=headers,
    )
    assert deleted_global.status_code == 204

    no_default = app_client.get("/api/complaints/report-cover/default?site=м—°мӮ°лҚ”мғө", headers=headers)
    assert no_default.status_code == 200
    assert no_default.json()["source_scope"] == "none"
    assert no_default.json()["company_name"] is None


def test_complaint_admin_record_grid_bulk_api(app_client: TestClient) -> None:
    headers = _owner_headers()
    created = app_client.post(
        "/api/complaints",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "building": "105лҸҷ",
            "unit_number": "1102нҳё",
            "description": "кҙҖлҰ¬ нғӯ л°°м№ҳмҲҳм • н…ҢмҠӨнҠё",
            "contact_phone": "010-1111-9999",
        },
    )
    assert created.status_code == 201
    complaint_id = created.json()["id"]

    added_event = app_client.post(
        f"/api/complaints/{complaint_id}/events",
        headers=headers,
        json={"event_type": "note", "note": "л°°м№ҳ мҲҳм • м „", "detail": {"step": 1}},
    )
    assert added_event.status_code == 201
    event_id = added_event.json()["id"]

    attachment = app_client.post(
        f"/api/complaints/{complaint_id}/attachments",
        headers=headers,
        data={"attachment_kind": "intake", "note": "кҙҖлҰ¬ нғӯ мІЁл¶Җ"},
        files={"file": ("admin-grid.txt", b"admin-grid", "text/plain")},
    )
    assert attachment.status_code == 201
    attachment_id = attachment.json()["id"]

    message = app_client.post(
        f"/api/complaints/{complaint_id}/messages",
        headers=headers,
        json={"body": "кҙҖлҰ¬ нғӯ л©”мӢңм§Җ"},
    )
    assert message.status_code == 201
    message_id = message.json()["id"]

    cost = app_client.post(
        f"/api/complaints/{complaint_id}/cost-items",
        headers=headers,
        json={"cost_category": "other", "item_name": "кҙҖлҰ¬ нғӯ л№„мҡ©", "quantity": 1, "unit_price": 5000},
    )
    assert cost.status_code == 201
    cost_item_id = cost.json()["id"]

    listed_cases = app_client.get(
        "/api/complaints/admin/records?site=м—°мӮ°лҚ”мғө&record_type=cases&limit=50&q=кҙҖлҰ¬ нғӯ",
        headers=headers,
    )
    assert listed_cases.status_code == 200
    listed_cases_body = listed_cases.json()
    assert listed_cases_body["record_type"] == "cases"
    assert listed_cases_body["total_count"] >= 1
    assert any(row["id"] == complaint_id for row in listed_cases_body["rows"])

    listed_events = app_client.get(
        "/api/complaints/admin/records?site=м—°мӮ°лҚ”мғө&record_type=events&limit=50&q=л°°м№ҳ мҲҳм • м „",
        headers=headers,
    )
    assert listed_events.status_code == 200
    assert listed_events.json()["total_count"] >= 1

    listed_attachments = app_client.get(
        "/api/complaints/admin/records?site=м—°мӮ°лҚ”мғө&record_type=attachments&limit=50&q=admin-grid.txt",
        headers=headers,
    )
    assert listed_attachments.status_code == 200
    assert listed_attachments.json()["total_count"] == 1

    listed_messages = app_client.get(
        "/api/complaints/admin/records?site=м—°мӮ°лҚ”мғө&record_type=messages&limit=50&q=кҙҖлҰ¬ нғӯ л©”мӢңм§Җ",
        headers=headers,
    )
    assert listed_messages.status_code == 200
    assert listed_messages.json()["total_count"] == 1

    listed_costs = app_client.get(
        "/api/complaints/admin/records?site=м—°мӮ°лҚ”мғө&record_type=cost_items&limit=50&q=кҙҖлҰ¬ нғӯ л№„мҡ©",
        headers=headers,
    )
    assert listed_costs.status_code == 200
    assert listed_costs.json()["total_count"] == 1

    updated_cases = app_client.post(
        "/api/complaints/admin/records/bulk-update",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "record_type": "cases",
            "rows": [
                {
                    "record_id": complaint_id,
                    "changes": {
                        "title": "кҙҖлҰ¬ нғӯ л°°м№ҳмҲҳм • мҷ„лЈҢ",
                        "assignee": "мқјкҙ„нҺём§‘л°ҳ",
                        "status": "assigned",
                    },
                }
            ],
        },
    )
    assert updated_cases.status_code == 200
    assert updated_cases.json()["updated_count"] == 1

    updated_events = app_client.post(
        "/api/complaints/admin/records/bulk-update",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "record_type": "events",
            "rows": [
                {
                    "record_id": event_id,
                    "changes": {
                        "event_type": "rework",
                        "note": "л°°м№ҳ мҲҳм • нӣ„",
                        "detail_json": "{\"step\": 2}",
                    },
                }
            ],
        },
    )
    assert updated_events.status_code == 200
    assert updated_events.json()["updated_count"] == 1

    case_detail = app_client.get(f"/api/complaints/{complaint_id}", headers=headers)
    assert case_detail.status_code == 200
    case_detail_body = case_detail.json()
    assert case_detail_body["case"]["title"] == "кҙҖлҰ¬ нғӯ л°°м№ҳмҲҳм • мҷ„лЈҢ"
    assert case_detail_body["case"]["assignee"] == "мқјкҙ„нҺём§‘л°ҳ"
    assert case_detail_body["case"]["status"] == "assigned"
    assert any(item["id"] == event_id and item["note"] == "л°°м№ҳ мҲҳм • нӣ„" for item in case_detail_body["events"])

    deleted_children = app_client.post(
        "/api/complaints/admin/records/bulk-delete",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "record_type": "attachments",
            "record_ids": [attachment_id],
        },
    )
    assert deleted_children.status_code == 200
    assert deleted_children.json()["deleted_count"] == 1

    deleted_messages = app_client.post(
        "/api/complaints/admin/records/bulk-delete",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "record_type": "messages",
            "record_ids": [message_id],
        },
    )
    assert deleted_messages.status_code == 200
    assert deleted_messages.json()["deleted_count"] == 1

    deleted_costs = app_client.post(
        "/api/complaints/admin/records/bulk-delete",
        headers=headers,
        json={
            "site": "м—°мӮ°лҚ”мғө",
            "record_type": "cost_items",
            "record_ids": [cost_item_id],
        },
    )
    assert deleted_costs.status_code == 200
    assert deleted_costs.json()["deleted_count"] == 1

    after_delete_detail = app_client.get(f"/api/complaints/{complaint_id}", headers=headers)
    assert after_delete_detail.status_code == 200
    after_delete_body = after_delete_detail.json()
    assert after_delete_body["attachments"] == []
    assert after_delete_body["messages"] == []
    assert after_delete_body["cost_items"] == []
